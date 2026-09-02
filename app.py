
import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
import os, tempfile, shutil, zipfile, re
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from supabase import create_client

st.set_page_config(
    page_title="Control Fases SFCO211",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)



# v19 · Estilo visual uniforme
st.markdown("""
<style>
/* Menú lateral azul más claro */
[data-testid="stSidebar"],
[data-testid="stSidebar"] > div:first-child {
    background: #6F91B3 !important;
}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span {
    color: #FFFFFF !important;
}

/* Elemento seleccionado del menú: contraste suave */
[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
    background: rgba(255,255,255,0.16) !important;
    border-radius: 8px !important;
}

/* Tamaño visual uniforme para gráficos */
[data-testid="stVegaLiteChart"],
[data-testid="stPlotlyChart"],
[data-testid="stPyplot"] {
    width: 100% !important;
    max-width: 100% !important;
}

/* Área de trabajo estable */
.block-container {
    max-width: 1850px !important;
    padding-left: 1.4rem !important;
    padding-right: 1.4rem !important;
}
</style>
""", unsafe_allow_html=True)

BASE = os.path.join(os.path.dirname(__file__), "CONTROL_FASES_SFCO211.xlsx")
LOGO = os.path.join(os.path.dirname(__file__), "logo_san_francisco.png")
PHASES = ["FASE1", "FASE2", "FASE3", "FASE4"]

@st.cache_resource(show_spinner=False)
def get_supabase():
    try:
        url = str(st.secrets["supabase"]["url"]).strip()
        key = str(st.secrets["supabase"]["service_role_key"]).strip()

        if not url.startswith("https://") or not url.endswith(".supabase.co"):
            raise ValueError("La URL de Supabase no tiene el formato esperado.")

        if not key.startswith("sb_secret_"):
            raise ValueError("La clave secreta no comienza con sb_secret_.")

        client = create_client(url, key)
        client.table("phase_updates").select("id").limit(1).execute()
        return client
    except Exception as e:
        st.session_state["supabase_error"] = str(e)
        return None

def db_ready():
    return get_supabase() is not None

def supabase_status():
    if db_ready():
        return True, "Conectada"
    return False, st.session_state.get("supabase_error", "No se pudo inicializar Supabase.")

def _db_select_all(table_name, select="*", order_by=None, desc=False, page_size=1000):
    """Lee TODOS los registros de una tabla Supabase mediante paginación.

    Supabase/PostgREST limita el número de filas devueltas por petición. Si no se
    pagina, una obra con muchas actualizaciones puede reconstruirse parcialmente y
    hacer que los porcentajes parezcan retroceder aunque los datos sí estén guardados.
    """
    client = get_supabase()
    if client is None:
        return []
    rows = []
    start = 0
    try:
        while True:
            query = client.table(table_name).select(select)
            if order_by:
                query = query.order(order_by, desc=bool(desc))
            res = query.range(start, start + page_size - 1).execute()
            batch = res.data or []
            rows.extend(batch)
            if len(batch) < page_size:
                break
            start += page_size
        return rows
    except Exception as e:
        st.session_state[f"{table_name}_read_error"] = str(e)
        return []


def db_phase_updates():
    # FUENTE PERSISTENTE PRINCIPAL: nunca leerla sin paginación.
    rows = _db_select_all(
        "phase_updates",
        select="*",
        order_by="id",
        desc=False,
        page_size=1000,
    )
    st.session_state["phase_updates_loaded_count"] = len(rows)
    return rows

def db_phase_update_history(limit=5000):
    """Historial inmutable de ediciones (disponible desde v41).
    La lectura se pagina para evitar perder registros antiguos al crecer la tabla.
    """
    rows = _db_select_all(
        "phase_update_history",
        select="*",
        order_by="changed_at",
        desc=True,
        page_size=1000,
    )
    return rows[:int(limit)] if limit else rows

def db_log_phase_update(phase, excel_row, activity, old_percent, new_percent, username):
    """Registra una edición en el historial sin bloquear el autoguardado si la tabla no existe."""
    client = get_supabase()
    if client is None:
        return False
    try:
        client.table("phase_update_history").insert({
            "phase": str(phase),
            "excel_row": int(excel_row),
            "activity": str(activity),
            "old_percent": None if old_percent is None else float(old_percent),
            "new_percent": float(new_percent),
            "changed_by": str(username),
            "changed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        }).execute()
        return True
    except Exception:
        return False

def db_weekly_detail_snapshot(snapshot_date=None, username=None):
    """Guarda una fotografía detallada de todas las celdas de avance.
    Se usa desde v41 para permitir recuperación exacta futura.
    """
    client = get_supabase()
    if client is None:
        return False, "Supabase no disponible"
    try:
        date_value = snapshot_date or datetime.now().date().isoformat()
        rows=[]
        # phase_updates contiene el estado persistente de las celdas modificadas.
        for r in db_phase_updates():
            rows.append({
                "snapshot_date": date_value,
                "phase": str(r.get("phase")),
                "excel_row": int(r.get("excel_row")),
                "activity": str(r.get("activity")),
                "percent": float(r.get("percent",0)),
                "captured_by": str(username or st.session_state.get("user_name","Usuario")),
            })
        if rows:
            # Borrar/recrear solo el mismo día para evitar duplicados parciales.
            try:
                client.table("weekly_detail_history").delete().eq("snapshot_date", date_value).execute()
            except Exception:
                pass
            for i in range(0,len(rows),500):
                client.table("weekly_detail_history").insert(rows[i:i+500]).execute()
        return True, f"Snapshot detallado {date_value}: {len(rows)} registros."
    except Exception as e:
        return False, str(e)

def db_upsert_phase_update(phase, excel_row, activity, percent, username):
    """Guarda una celda en Supabase y confirma que quedó persistida."""
    client = get_supabase()
    if client is None:
        return False, st.session_state.get("supabase_error", "Supabase no está disponible.")
    payload = {
        "phase": str(phase),
        "excel_row": int(excel_row),
        "activity": str(activity),
        "percent": float(percent),
        "updated_by": str(username),
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }
    try:
        old_percent = None
        try:
            before = (
                client.table("phase_updates")
                .select("percent")
                .eq("phase", str(phase))
                .eq("excel_row", int(excel_row))
                .eq("activity", str(activity))
                .limit(1)
                .execute()
            )
            if before.data:
                old_percent = float(before.data[0].get("percent"))
        except Exception:
            old_percent = None

        client.table("phase_updates").upsert(
            payload, on_conflict="phase,excel_row,activity"
        ).execute()
        check = (
            client.table("phase_updates")
            .select("percent,updated_by,updated_at")
            .eq("phase", str(phase))
            .eq("excel_row", int(excel_row))
            .eq("activity", str(activity))
            .limit(1)
            .execute()
        )
        rows = check.data or []
        if not rows:
            return False, "Supabase no devolvió el registro después de guardar."
        saved = float(rows[0].get("percent", -999))
        if abs(saved - float(percent)) > 1e-6:
            return False, f"Verificación inconsistente: se pidió {percent}% y la base devolvió {saved}%."
        # Historial inmutable: desde v41 cada cambio queda registrado además del estado actual.
        if old_percent is None or abs(float(old_percent) - float(percent)) > 1e-6:
            db_log_phase_update(phase, excel_row, activity, old_percent, percent, username)
        return True, rows[0].get("updated_at") or payload["updated_at"]
    except Exception as e:
        return False, str(e)


def workbook_recovery_payload(path, username):
    """Convierte FASE1–FASE4 del Excel actual en registros 0–100 para phase_updates.

    Se usan las mismas claves persistentes de la aplicación:
    (phase, excel_row, activity). No se incluye la columna calculada
    "% Avance Real Depto" porque se recalcula desde las partidas.
    """
    wb = load_workbook(path, read_only=True, data_only=False)
    payload = []
    counts = {}
    stamp = datetime.now().astimezone().isoformat(timespec="seconds")
    try:
        for phase in PHASES:
            if phase not in wb.sheetnames:
                raise ValueError(f"El archivo no contiene la hoja {phase}.")
            ws = wb[phase]
            headers = [c.value for c in ws[1]]
            if "% Avance Real Depto" not in headers:
                raise ValueError(f"{phase}: falta la columna % Avance Real Depto.")
            progress_idx = headers.index("% Avance Real Depto") + 1
            activity_cols = []
            for col_idx in range(5, progress_idx):
                activity = headers[col_idx - 1]
                if activity is not None and str(activity).strip():
                    activity_cols.append((col_idx, str(activity)))

            phase_count = 0
            for excel_row in range(2, ws.max_row + 1):
                # Solo filas reales de departamentos/sectores: Fase, Piso, Torre y Departamento.
                id_values = [ws.cell(excel_row, c).value for c in range(1, 5)]
                if not any(v is not None and str(v).strip() != "" for v in id_values):
                    continue
                if ws.cell(excel_row, 4).value in (None, ""):
                    continue

                for col_idx, activity in activity_cols:
                    raw = ws.cell(excel_row, col_idx).value
                    # En una base de recuperación, una celda vacía se interpreta como 0%.
                    pct = to_percent(0 if raw in (None, "") else raw, phase)
                    pct = max(0.0, min(100.0, float(pct)))
                    payload.append({
                        "phase": phase,
                        "excel_row": int(excel_row),
                        "activity": activity,
                        "percent": round(pct, 6),
                        "updated_by": f"RECUPERACION EXCEL · {username}",
                        "updated_at": stamp,
                    })
                    phase_count += 1
            counts[phase] = phase_count
    finally:
        wb.close()
    return payload, counts


def db_replace_from_recovery_workbook(path, username):
    """Sobrescribe/crea el estado oficial de cada celda del Excel en Supabase.

    No borra la tabla antes de escribir: así una interrupción de red no puede dejar
    phase_updates vacía. Las claves existentes se sobrescriben por la restricción
    única (phase, excel_row, activity); por tanto no se duplican registros.
    """
    client = get_supabase()
    if client is None:
        return False, st.session_state.get("supabase_error", "Supabase no está disponible."), {}

    try:
        payload, counts = workbook_recovery_payload(path, username)
        if not payload:
            return False, "El Excel no contiene partidas válidas para sincronizar.", counts

        # Escribir primero y verificar después. Nunca vaciar la tabla antes del upsert.
        batch_size = 400
        for i in range(0, len(payload), batch_size):
            client.table("phase_updates").upsert(
                payload[i:i + batch_size],
                on_conflict="phase,excel_row,activity"
            ).execute()

        # Verificación completa usando lectura paginada (protección v40+).
        online = db_phase_updates()
        online_map = {}
        for r in online:
            try:
                key = (str(r.get("phase")), int(r.get("excel_row")), str(r.get("activity")))
                online_map[key] = float(r.get("percent", 0))
            except Exception:
                continue

        mismatches = []
        for r in payload:
            key = (r["phase"], int(r["excel_row"]), r["activity"])
            db_value = online_map.get(key)
            if db_value is None or abs(float(db_value) - float(r["percent"])) > 1e-6:
                mismatches.append((key, r["percent"], db_value))
                if len(mismatches) >= 10:
                    break

        if mismatches:
            return False, (
                "La carga terminó, pero la verificación encontró diferencias. "
                f"Primeras diferencias: {mismatches[:3]}"
            ), counts

        return True, f"Base online sincronizada y verificada: {len(payload)} celdas de avance.", counts
    except Exception as e:
        return False, str(e), {}

def set_save_status(ok, message, phase=None, excel_row=None, activity=None):
    st.session_state["last_save_status"] = {
        "ok": bool(ok),
        "message": str(message),
        "phase": phase,
        "excel_row": excel_row,
        "activity": activity,
        "local_time": datetime.now().astimezone().strftime("%d-%m-%Y %H:%M:%S"),
    }

def show_sync_diagnostics():
    """Indicador simple para confirmar que la reconstrucción online fue completa."""
    count = st.session_state.get("phase_updates_loaded_count")
    err = st.session_state.get("phase_updates_read_error")
    if err:
        st.error(f"❌ Error leyendo avances online: {err}")
    elif count is not None:
        st.caption(f"☁️ Registros de avance cargados desde Supabase: {count:,}".replace(",", "."))

def show_save_status():
    info = st.session_state.get("last_save_status")
    if not info:
        return
    prefix = "✅ Autoguardado confirmado" if info.get("ok") else "❌ Error de autoguardado"
    detail = f"{prefix} · {info.get('local_time','')}"
    if info.get("phase"):
        detail += f" · {str(info['phase']).replace('FASE','Fase ')}"
    if info.get("activity"):
        detail += f" · {info['activity']}"
    if info.get("ok"):
        st.success(detail)
    else:
        st.error(detail + f" · {info.get('message','')}")

def db_weekly_history():
    # También se pagina para que el histórico siga siendo completo con el tiempo.
    return _db_select_all(
        "weekly_history",
        select="*",
        order_by="update_date",
        desc=False,
        page_size=1000,
    )

def db_latest_weekly_snapshot():
    """Devuelve el último corte oficial directamente desde Supabase, sin cache.
    Esta función es la fuente maestra para Dashboard, Ruta Crítica y Gantt.
    """
    client = get_supabase()
    if client is None:
        return None
    try:
        res = (
            client.table("weekly_history")
            .select("*")
            .order("update_date", desc=True)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        return rows[0] if rows else None
    except Exception:
        return None

def db_register_weekly(summaries, general, username):
    client = get_supabase()
    if client is None:
        return False, "Base de datos no configurada."
    date_value = datetime.now().date().isoformat()
    payload = {
        "update_date": date_value,
        "general": float(general),
        "fase1": float(summaries["FASE1"]),
        "fase2": float(summaries["FASE2"]),
        "fase3": float(summaries["FASE3"]),
        "fase4": float(summaries["FASE4"]),
        "updated_by": str(username),
    }
    try:
        client.table("weekly_history").upsert(
            payload, on_conflict="update_date"
        ).execute()
        snap_ok, snap_msg = db_weekly_detail_snapshot(date_value, username)
        extra = f" Snapshot detallado: {snap_msg}" if snap_ok else ""
        return True, f"Actualización semanal guardada en línea: {date_value}." + extra
    except Exception as e:
        return False, f"No se pudo guardar en la base de datos: {e}"


st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#f4f7fb;}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#071f36 0%,#0a3b63 100%);}
[data-testid="stSidebar"] *{color:#fff;}
.block-container{padding-top:1.1rem;max-width:1700px;}
.sf-title{font-size:30px;font-weight:850;color:#10233d;letter-spacing:-.4px;margin-bottom:1px}
.sf-sub{font-size:14px;color:#708096;margin-bottom:22px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:15px;padding:18px;box-shadow:0 4px 16px rgba(15,35,60,.04)}
.klabel{font-size:11px;font-weight:800;color:#77849a;text-transform:uppercase}
.kvalue{font-size:30px;font-weight:850;color:#10233d;margin-top:5px}
.good{color:#179b59}.warn{color:#e79b13}.bad{color:#df4545}.blue{color:#1678e7}
.section{font-size:17px;font-weight:850;color:#1b2d47;margin:14px 0 10px}
.phase-card{background:#fff;border:1px solid #dfe6ef;border-radius:16px;padding:16px 18px;
box-shadow:0 4px 14px rgba(15,35,60,.05);margin-bottom:8px}
.phase-name{font-size:12px;font-weight:800;color:#6f7d92;text-transform:uppercase}
.phase-pct{font-size:35px;font-weight:900;color:#10233d;line-height:1.1;margin-top:4px}
.general-box{background:linear-gradient(135deg,#102f50 0%,#1769aa 100%);border-radius:17px;
padding:19px 22px;color:white;box-shadow:0 5px 18px rgba(15,35,60,.12)}
.general-label{font-size:12px;font-weight:800;opacity:.82;text-transform:uppercase}
.general-pct{font-size:39px;font-weight:900;line-height:1.05;margin-top:4px}

.stButton>button,.stDownloadButton>button{border-radius:9px;font-weight:750}
div[data-testid="stDataFrame"],div[data-testid="stDataEditor"]{
 background:#fff;border-radius:12px;border:1px solid #e2e8f0;padding:5px;
}
div[data-testid="stDataFrame"] {font-size:12px;}

@media (max-width: 768px) {
  .block-container {padding: .65rem .65rem 1.5rem .65rem;}
  .sf-title {font-size:22px !important; line-height:1.1;}
  .sf-sub {font-size:12px !important; margin-bottom:12px;}
  .card {padding:12px; min-height:96px;}
  .kvalue {font-size:24px !important;}
  .phase-pct {font-size:27px !important;}
  .general-pct {font-size:31px !important;}
  div[data-testid="stHorizontalBlock"] {gap:.45rem;}
  [data-testid="stSidebar"] img {max-height:150px; object-fit:contain;}
  .stButton>button,.stDownloadButton>button {
      min-height:44px;
      font-size:15px;
  }
}
</style>
""", unsafe_allow_html=True)

APP_VERSION = "v46"

# Al cambiar de versión, reemplazar cualquier copia temporal antigua por
# la plantilla depurada incluida en esta versión. Los avances reales se
# vuelven a aplicar desde Supabase.
if "workbook_path" not in st.session_state:
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(BASE, p)
    st.session_state.workbook_path = p
    st.session_state.app_version = APP_VERSION
elif st.session_state.get("app_version") != APP_VERSION:
    shutil.copy2(BASE, st.session_state.workbook_path)
    st.session_state.app_version = APP_VERSION

def get_users():
    """
    Usuarios almacenados en Streamlit Secrets.
    Formato:
    [users.admin]
    password = "..."
    name = "..."
    role = "Administrador"
    """
    try:
        users = st.secrets.get("users", {})
        return users
    except Exception:
        return {}

def login_screen():
    users = get_users()

    if not users:
        st.error("No hay usuarios configurados todavía.")
        st.info(
            "Configura los usuarios en Streamlit → App settings → Secrets. "
            "Usa el formato incluido en el archivo secrets.toml.example."
        )
        st.stop()

    st.markdown("## 🔐 Ingreso al Control de Obra")
    with st.form("login_form"):
        username = st.text_input("Usuario")
        password = st.text_input("Clave", type="password")
        submitted = st.form_submit_button("INGRESAR", use_container_width=True)

    if submitted:
        if username in users and str(users[username].get("password", "")) == password:
            st.session_state.authenticated = True
            st.session_state.username = username
            st.session_state.user_name = users[username].get("name", username)
            st.session_state.user_role = users[username].get("role", "Usuario")
            st.session_state.user_phase = users[username].get("phase", "ALL")
            st.rerun()
        else:
            st.error("Usuario o clave incorrectos.")

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    if os.path.exists(LOGO):
        c1, c2, c3 = st.columns([1, 2, 1])
        with c2:
            st.image(LOGO, use_container_width=True)
    login_screen()
    st.stop()

def phase_scale(sheet_name):
    # El archivo usa FASE1 en 0-100 y FASE2/3/4 en 0-1.
    return 100.0 if sheet_name == "FASE1" else 1.0

def to_percent(value, phase):
    try:
        if value is None or value == "":
            return 0.0
        x = float(value)
        return x if phase_scale(phase) == 100.0 else x * 100.0
    except:
        return 0.0

def from_percent(value, phase):
    x = max(0.0, min(100.0, float(value)))
    return x if phase_scale(phase) == 100.0 else x / 100.0

@st.cache_data(show_spinner=False)
def get_sheet_names(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    names = wb.sheetnames
    wb.close()
    return names

def load_phase(path, phase):
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb[phase]
    headers = [c.value for c in ws[1]]
    rows, excel_rows = [], []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in vals):
            continue
        rows.append(vals)
        excel_rows.append(r)
    wb.close()
    df = pd.DataFrame(rows, columns=headers)
    df["_excel_row"] = excel_rows

    for upd in db_phase_updates():
        try:
            if upd.get("phase") != phase:
                continue
            excel_row = int(upd.get("excel_row"))
            activity = upd.get("activity")
            pct = float(upd.get("percent"))
            if activity in df.columns:
                mask = df["_excel_row"] == excel_row
                if mask.any():
                    df.loc[mask, activity] = from_percent(pct, phase)
        except Exception:
            pass
    return df

def phase_editor_df(path, phase):
    """
    Devuelve la misma tabla que se usa en "Fases completas".
    Todas las partidas se muestran en 0-100 y el avance real se recalcula
    directamente desde las partidas.
    """
    df = load_phase(path, phase).copy()
    id_cols = ["Fase", "Piso", "Torre", "Departamento", "_excel_row"]
    progress_col = "% Avance Real Depto"

    activity_cols = [
        c for c in df.columns
        if c not in id_cols and c != progress_col
    ]

    for col in activity_cols:
        vals = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
        if phase_scale(phase) == 1.0:
            vals = vals * 100.0
        df[col] = vals.round(1)

    if activity_cols:
        df[progress_col] = df[activity_cols].mean(axis=1).round(1)
    else:
        df[progress_col] = 0.0

    return df

def phase_numeric_percent_df(path, phase):
    """
    FUENTE ÚNICA PARA DASHBOARD Y GRÁFICOS.
    Usa exactamente las partidas visibles en "Fases completas" y calcula
    el % Avance Real Depto directamente desde esas partidas, sin depender
    de las fórmulas guardadas/caché del Excel.
    """
    df = phase_editor_df(path, phase).copy()

    id_cols = ["Fase", "Piso", "Torre", "Departamento", "_excel_row"]
    progress_col = "% Avance Real Depto"

    activity_cols = [
        c for c in df.columns
        if c not in id_cols and c != progress_col
    ]

    # Convertir todas las partidas a números 0-100.
    for col in activity_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Recalcular avance real del departamento desde las partidas.
    if activity_cols:
        df[progress_col] = df[activity_cols].mean(axis=1).round(2)
    else:
        df[progress_col] = 0.0

    return df

def phase_summary(path, phase):
    """
    AVANCE OFICIAL DE FASE.
    Replica el criterio de la hoja RESUMEN de la planilla:
    - toma las partidas de Fases completas,
    - calcula el promedio de cada partida por piso,
    - considera los pisos 1 a 9 con el mismo peso,
    - luego obtiene el promedio general de todas las partidas.
    """
    df = phase_numeric_percent_df(path, phase)

    if df.empty or "Piso" not in df.columns:
        return 0.0

    skip = {
        "Fase", "Piso", "Torre", "Departamento",
        "% Avance Real Depto", "_excel_row"
    }
    activity_cols = [c for c in df.columns if c not in skip]

    if not activity_cols:
        return 0.0

    # La planilla RESUMEN contempla Piso 1 a Piso 9.
    official_floors = list(range(1, 10))
    activity_results = []

    for activity in activity_cols:
        floor_results = []

        for piso in official_floors:
            piso_rows = df[pd.to_numeric(df["Piso"], errors="coerce") == piso]

            if piso_rows.empty:
                floor_avg = 0.0
            else:
                vals = pd.to_numeric(
                    piso_rows[activity],
                    errors="coerce"
                ).fillna(0.0)
                floor_avg = float(vals.mean()) if len(vals) else 0.0

            floor_results.append(floor_avg)

        activity_results.append(
            sum(floor_results) / len(official_floors)
        )

    return float(round(
        sum(activity_results) / len(activity_results)
    )) if activity_results else 0.0

def phase_by_floor(path, phase):
    df = phase_numeric_percent_df(path, phase)
    if "Piso" not in df.columns or "% Avance Real Depto" not in df.columns:
        return pd.DataFrame(columns=["Piso","Avance (%)"])
    out = (
        df.groupby("Piso", dropna=True)["% Avance Real Depto"]
        .mean()
        .reset_index()
        .rename(columns={"% Avance Real Depto":"Avance (%)"})
    )
    out["Avance (%)"] = out["Avance (%)"].round(1)
    return out

def phase_by_tower(path, phase):
    df = phase_numeric_percent_df(path, phase)
    if "Torre" not in df.columns or "% Avance Real Depto" not in df.columns:
        return pd.DataFrame(columns=["Torre","Avance (%)"])
    out = (
        df.groupby("Torre", dropna=True)["% Avance Real Depto"]
        .mean()
        .reset_index()
        .rename(columns={"% Avance Real Depto":"Avance (%)"})
    )
    out["Avance (%)"] = out["Avance (%)"].round(1)
    return out

def phase_activity_summary(path, phase):
    df = phase_numeric_percent_df(path, phase)
    skip = {"Fase","Piso","Torre","Departamento","% Avance Real Depto","_excel_row"}
    rows=[]
    for col in df.columns:
        if col not in skip:
            vals = pd.to_numeric(df[col], errors="coerce")
            if vals.notna().any():
                rows.append({"Partida":col, "Avance (%)":round(float(vals.fillna(0).mean()),1)})
    return pd.DataFrame(rows).sort_values("Avance (%)", ascending=False)

def phase_activity_floor_matrix(path, phase):
    """Matriz de partidas x pisos, alimentada directamente por Fases completas."""
    df = phase_numeric_percent_df(path, phase)
    skip = {"Fase","Piso","Torre","Departamento","% Avance Real Depto","_excel_row"}
    activities = [c for c in df.columns if c not in skip]
    rows = []
    piso_num = pd.to_numeric(df.get("Piso"), errors="coerce")
    for activity in activities:
        rec = {"Fase": phase.replace("FASE", "FASE "), "Partida": activity}
        vals_all = []
        for piso in range(1,10):
            vals = pd.to_numeric(df.loc[piso_num == piso, activity], errors="coerce").fillna(0.0)
            avg = float(vals.mean()) if len(vals) else 0.0
            avg = round(avg, 1)
            rec[f"Piso {piso}"] = avg
            vals_all.append(avg)
        real = round(sum(vals_all)/9, 1) if vals_all else 0.0
        rec["% Avance Real"] = real
        if real >= 100:
            rec["Estado"] = "✅ Completado"
        elif real >= 50:
            rec["Estado"] = "🟡 En progreso"
        elif real > 0:
            rec["Estado"] = "🟠 En riesgo"
        else:
            rec["Estado"] = "⚪ No iniciado"
        rows.append(rec)
    return pd.DataFrame(rows)


# Ruta Crítica oficial definida por el usuario (imagen de referencia 24-08-2026).
# Cada etiqueta visible se enlaza con la columna real de la fase para conservar actualización automática.
CRITICAL_ROUTE_COMPONENTS = {
    "FASE1": [
        ("Trazado tabique/Yeso/auxiliares vent. Y puertas", "Trazado tabique/Yeso/auxiliares vent. Y puertas"),
        ("Grada Buque (9cm)", "Grada Buque (9cm)"),
        ("Rasgo ventana - terrazas", "Rasgo ventana - terrazas"),
        ("Yeso Muro", "Yeso Muro"),
        ("Yeso Cielo", "Yeso Cielo"),
        ("Sanitario vertical y ramal", "Sanitario vertical y ramal"),
        ("Tabique 1° Cara", "Tabique 1° Cara"),
        ("Tabique 2° cara", "Tabique 2° cara "),
    ],
    "FASE2": [
        ("Huinchas tabiques", "Huinchas tabiques"),
        ("Instalación Marco Puerta Acceso", "Instalación Marco Puerta Acceso"),
        ("Marcos Puertas Interiores", "Marcos Puertas Interiores"),
        ("Impermeabilizacion Baños y cocina", "Impermeabilizacion Baños"),
        ("Instalación de Receptaculo", "Instalacion de tina/Receptaculo"),
        ("Protección Tina/receptaculo", "Protección Tina/receptaculo"),
        ("Instalación Ventana", "Instalación Ventana"),
        ("Ceramica Pisos", "Ceramica Pisos"),
        ("Ceramica Muros", "Ceramica Muros"),
        ("Cerámica de Terrazas", "Cerámica de Terrazas"),
    ],
    "FASE3": [
        ("Instalación piernas de closet", "Instalación piernas de closet"),
        ("Instalación Cornisas", "Instalación Cornisas"),
        ("Puertas interiores", "Puertas interiores"),
        ("Puertas de acceso y pasillo", "Puertas de acceso y pasillo"),
        ("Cerradura Puertas", "Cerradura Puertas "),
        ("Cableado Eléctrico y CCDD", "Cableado Eléctrico y CCDD"),
        ("Empaste Muro Cielo", "Empaste Muro Cielo"),
        ("Instalación de Baranda de Cristal", "Instalación de Baranda de Crital"),
        ("Losalin Cielo Departamento", "Losalin Cielo Departamento "),
        ("Instalacion Muebles de cocina (Base)", "Instalacion de Muebles de cocina (Base)"),
        ("Instalacion Muebles de cocina (Aereo)", "Instalacion Muebles de cocina (Aereo)"),
        ("Instalación cubiertas cocina", "Instalación cubiertas cocina"),
    ],
    "FASE4": [
        ("Instalacion de Artefactos Sanitarios.", "Instalacion de Artefactos Sanitarios."),
        ("Instalación de Shower Door", "Instalación de Shower Door y Barra Cortina"),
        ("Kit de cocina (Horno-Encimera -Campana - Lavaplato)", "Kit de cocina (Horno-Encimera -Campana - Lavaplato)"),
        ("Griferia.", "Griferia."),
        ("Piso Flotante + cubrejunta + junquillo", "Piso Flotante + cubrejunta + junquillo"),
        ("2da Mano de Pintura", "2° mano de pintura. (puertas, , cielo de baño, )"),
        ("Instalación de Papel Mural", "Instalación Papel Mural"),
        ("Topes en piso y Accesorios", "Topes en piso y Accesorios"),
    ],
}

def critical_route_matrix(path, phase):
    """Matriz oficial de Ruta Crítica según selección indicada por el usuario."""
    df = phase_numeric_percent_df(path, phase)
    if df.empty:
        return pd.DataFrame()
    piso_num = pd.to_numeric(df.get("Piso"), errors="coerce")
    rows = []
    for visible_name, source_col in CRITICAL_ROUTE_COMPONENTS.get(phase, []):
        if source_col not in df.columns:
            continue
        rec = {"Fase": phase.replace("FASE", "FASE "), "Partida": visible_name}
        floor_vals = []
        for piso in range(1, 10):
            vals = pd.to_numeric(df.loc[piso_num == piso, source_col], errors="coerce").fillna(0.0)
            avg = round(float(vals.mean()) if len(vals) else 0.0, 1)
            rec[f"Piso {piso}"] = avg
            floor_vals.append(avg)
        real = round(sum(floor_vals) / 9, 1) if floor_vals else 0.0
        rec["% Avance Real"] = real
        if real >= 100:
            rec["Estado"] = "✅ Completado"
        elif real >= 50:
            rec["Estado"] = "🟡 En progreso"
        elif real > 0:
            rec["Estado"] = "🟠 En riesgo"
        else:
            rec["Estado"] = "⚪ No iniciado"
        rows.append(rec)
    return pd.DataFrame(rows)

def _pct_style(v):
    try:
        x = float(v)
    except Exception:
        return ""
    if x >= 99.999:
        return "background-color:#22c55e;color:#062b13;font-weight:800"
    if x >= 50:
        return "background-color:#fde047;color:#3d3000;font-weight:700"
    return "background-color:#fb923c;color:#421900;font-weight:700"

def style_percent_df(df, percent_columns=None):
    """Semáforo de 3 tonos: verde=100%, amarillo=50-99%, naranjo=0-49%."""
    if percent_columns is None:
        percent_columns = [c for c in df.columns if "%" in str(c) or str(c).lower().startswith("piso ")]
    percent_columns = [c for c in percent_columns if c in df.columns]
    styler = df.style
    if percent_columns:
        styler = styler.map(_pct_style, subset=percent_columns)
        styler = styler.format({c:"{:.1f}%" for c in percent_columns}, na_rep="")
    return styler

# Programación oficial Gantt – Alternativa 2.
# Orden fijo: Fase 1 Piso 1-9, Fase 2 Piso 1-9, Fase 3 Piso 1-9 y Fase 4 Piso 1-9.
# Cada piso termina el último día hábil del mismo mes en que inicia.
GANTT_PLAN = {
    # Fase 1 confirmada desde julio de 2026, un piso por mes.
    "FASE1": [(f"Piso {i}", f"{y:04d}-{m:02d}-01") for i,(y,m) in enumerate([
        (2026,7),(2026,8),(2026,9),(2026,10),(2026,11),(2026,12),(2027,1),(2027,2),(2027,3)
    ], start=1)],
    # Fases 2 a 4 mantienen sus inicios correlativos confirmados; solo cambia el término al último día hábil del mes.
    "FASE2": [(f"Piso {i}", d) for i,d in enumerate([
        "2026-10-04","2026-11-04","2026-12-04","2027-01-04","2027-02-03","2027-03-06","2027-04-06","2027-05-04","2027-06-04"
    ], start=1)],
    "FASE3": [(f"Piso {i}", d) for i,d in enumerate([
        "2026-11-04","2026-12-04","2027-01-04","2027-02-03","2027-03-06","2027-04-06","2027-05-04","2027-06-04","2027-07-04"
    ], start=1)],
    "FASE4": [(f"Piso {i}", d) for i,d in enumerate([
        "2026-12-04","2027-01-04","2027-02-03","2027-03-06","2027-04-06","2027-05-04","2027-06-04","2027-07-04","2027-08-02"
    ], start=1)],
}

def _last_business_day_same_month(start):
    """Último lunes-viernes del mes de `start`."""
    start = pd.Timestamp(start).normalize()
    month_end = start + pd.offsets.MonthEnd(0)
    while month_end.weekday() >= 5:  # 5=sábado, 6=domingo
        month_end -= pd.Timedelta(days=1)
    return month_end.normalize()

def build_gantt_df(path, phases, review_date=None):
    """Construye la Gantt mensual y calcula atraso dinámico.

    Alternativa 2:
    - Cada piso termina el último día hábil del mismo mes en que inicia.
    - Orden fijo por fase y piso.
    - 100% completado: 0 días de atraso.
    - Incompleto después del término planificado: atraso en días calendario.
    """
    if review_date is None:
        try:
            from zoneinfo import ZoneInfo
            review_date = pd.Timestamp.now(tz=ZoneInfo("America/Santiago")).tz_localize(None).normalize()
        except Exception:
            review_date = pd.Timestamp.today().normalize()
    else:
        review_date = pd.Timestamp(review_date).normalize()

    rows=[]
    for ph in phases:
        pf = phase_by_floor(path, ph)
        floor_progress = {}
        if not pf.empty:
            for _, rr in pf.iterrows():
                try:
                    floor_progress[int(float(rr["Piso"]))] = float(rr["Avance (%)"])
                except Exception:
                    pass
        for floor_name, start_txt in GANTT_PLAN[ph]:
            piso_num = int(floor_name.split()[-1])
            start = pd.to_datetime(start_txt).normalize()
            finish = _last_business_day_same_month(start)
            business_days = len(pd.bdate_range(start=start, end=finish))
            progress = round(floor_progress.get(piso_num,0.0),1)
            delay_days = 0 if progress >= 100 else max(0, int((review_date - finish).days))
            if progress >= 100:
                schedule_status = "Completado"
            elif delay_days > 0:
                schedule_status = "Atrasado"
            elif review_date < start:
                schedule_status = "No iniciado"
            else:
                schedule_status = "En plazo"
            rows.append({
                "Fase": ph.replace("FASE","Fase "), "Piso": floor_name,
                "Tarea": f'{ph.replace("FASE","Fase ")} · {floor_name}',
                "Inicio": start, "Término": finish,
                "Días": int(business_days), "Progreso (%)": progress,
                "Días atraso": delay_days, "Estado plazo": schedule_status
            })
    return pd.DataFrame(rows)

def recalc_department_progress(ws, row, phase):
    first_activity = 5
    last_activity = ws.max_column - 1
    vals = []
    for c in range(first_activity, last_activity + 1):
        v = ws.cell(row, c).value
        try:
            if v is not None and v != "":
                vals.append(float(v))
        except:
            pass
    result = sum(vals) / len(vals) if vals else 0
    ws.cell(row, ws.max_column).value = result

def save_department(path, phase, excel_row, activity_values):
    """Guarda únicamente cambios confirmados; Supabase es la fuente persistente."""
    wb = load_workbook(path)
    ws = wb[phase]
    headers = [c.value for c in ws[1]]
    header_to_col = {str(h): i+1 for i,h in enumerate(headers) if h is not None}
    username = st.session_state.get("user_name", st.session_state.get("username", "Usuario"))
    saved = 0
    errors = []
    for activity, pct in activity_values.items():
        col = header_to_col.get(activity)
        if not col:
            continue
        ok, msg = db_upsert_phase_update(phase, excel_row, activity, pct, username)
        if not ok:
            errors.append(f"{activity}: {msg}")
            continue
        ws.cell(excel_row, col).value = from_percent(pct, phase)
        saved += 1
        set_save_status(True, msg, phase, excel_row, activity)

    if saved:
        recalc_department_progress(ws, excel_row, phase)
        wb.save(path)
    wb.close()
    st.cache_data.clear()
    if errors:
        set_save_status(False, " | ".join(errors), phase, excel_row)
    return saved, errors

def save_full_phase(path, phase, edited_df, original_df):
    """
    Guarda solamente las celdas modificadas de una fase completa.
    Los porcentajes visibles 0–100 vuelven a la escala interna correcta del Excel.
    """
    wb = load_workbook(path)
    ws = wb[phase]

    id_cols = ["Fase", "Piso", "Torre", "Departamento", "_excel_row"]
    headers = [c.value for c in ws[1]]
    header_to_col = {str(h): i+1 for i,h in enumerate(headers) if h is not None}

    changed_rows = set()
    changed_cells = 0

    for idx in edited_df.index:
        excel_row = int(original_df.loc[idx, "_excel_row"])

        for col in edited_df.columns:
            if col in id_cols or col == "% Avance Real Depto":
                continue

            new_val = edited_df.loc[idx, col]
            old_val = original_df.loc[idx, col]

            # Compare numerically where possible.
            try:
                new_num = float(new_val)
            except:
                new_num = 0.0
            try:
                old_num = float(old_val)
            except:
                old_num = 0.0

            if abs(new_num - old_num) < 1e-9:
                continue

            excel_col = header_to_col.get(str(col))
            if not excel_col:
                continue

            username = st.session_state.get("user_name", st.session_state.get("username", "Usuario"))
            ok, msg = db_upsert_phase_update(phase, excel_row, col, new_num, username)
            if not ok:
                wb.close()
                set_save_status(False, msg, phase, excel_row, col)
                raise RuntimeError(f"No se pudo guardar {phase} fila {excel_row} / {col}: {msg}")
            ws.cell(excel_row, excel_col).value = from_percent(new_num, phase)
            set_save_status(True, msg, phase, excel_row, col)
            changed_rows.add(excel_row)
            changed_cells += 1

    for excel_row in changed_rows:
        recalc_department_progress(ws, excel_row, phase)

    wb.save(path)
    wb.close()
    # Forzar refresco inmediato de Dashboard, gráficos, Gantt y editores.
    st.cache_data.clear()
    return changed_cells, len(changed_rows)


def _xlsx_col_number(cell_ref):
    letters = "".join(ch for ch in str(cell_ref) if ch.isalpha())
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n

def _xlsx_col_letter(n):
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out

def build_formatted_export():
    """
    Exporta directamente desde la plantilla XLSX depurada de la aplicación.
    No reabre ni vuelve a guardar el libro con openpyxl: modifica únicamente
    los valores numéricos de las partidas dentro del paquete XLSX.
    Esto conserva el formato y evita que Excel tenga que reparar el archivo.
    """
    with open(BASE, "rb") as f:
        base_bytes = f.read()

    updates = db_phase_updates()
    source = BytesIO(base_bytes)
    output = BytesIO()

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel_doc = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg_rel = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(source, "r") as zin:
        names = set(zin.namelist())

        # Shared strings para interpretar los encabezados de FASE1–FASE4.
        shared = []
        if "xl/sharedStrings.xml" in names:
            ss_root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
            for si in ss_root.findall(f"{{{ns_main}}}si"):
                shared.append("".join(t.text or "" for t in si.iter(f"{{{ns_main}}}t")))

        workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

        rel_targets = {}
        for rel in rel_root:
            rel_targets[rel.attrib.get("Id")] = rel.attrib.get("Target", "")

        sheet_paths = {}
        sheets_node = workbook_root.find(f"{{{ns_main}}}sheets")
        if sheets_node is not None:
            for sh in sheets_node:
                name = sh.attrib.get("name")
                rid = sh.attrib.get(f"{{{ns_rel_doc}}}id")
                target = rel_targets.get(rid, "")
                if target.startswith("/"):
                    path = target.lstrip("/")
                else:
                    path = os.path.normpath(os.path.join("xl", target)).replace("\\\\", "/")
                sheet_paths[name] = path

        # Preparar modificaciones por XML de hoja.
        per_sheet = {}
        for upd in updates:
            try:
                phase = str(upd.get("phase", "")).upper()
                if phase not in PHASES or phase not in sheet_paths:
                    continue
                per_sheet.setdefault(phase, []).append(upd)
            except Exception:
                continue

        replacements = {}

        for phase, phase_updates in per_sheet.items():
            path = sheet_paths[phase]
            if path not in names:
                continue

            root = ET.fromstring(zin.read(path))
            sheet_data = root.find(f"{{{ns_main}}}sheetData")
            if sheet_data is None:
                continue

            rows = {
                int(r.attrib.get("r", "0")): r
                for r in sheet_data.findall(f"{{{ns_main}}}row")
            }

            # Leer encabezados desde fila 1.
            header_to_col = {}
            header_row = rows.get(1)
            if header_row is not None:
                for c in header_row.findall(f"{{{ns_main}}}c"):
                    ref = c.attrib.get("r", "")
                    col_num = _xlsx_col_number(ref)
                    ctype = c.attrib.get("t")
                    value = ""
                    if ctype == "inlineStr":
                        is_node = c.find(f"{{{ns_main}}}is")
                        if is_node is not None:
                            value = "".join(t.text or "" for t in is_node.iter(f"{{{ns_main}}}t"))
                    else:
                        v = c.find(f"{{{ns_main}}}v")
                        if v is not None and v.text is not None:
                            if ctype == "s":
                                try:
                                    value = shared[int(v.text)]
                                except Exception:
                                    value = ""
                            else:
                                value = v.text
                    if value:
                        header_to_col[str(value)] = col_num

            for upd in phase_updates:
                try:
                    row_num = int(upd.get("excel_row"))
                    activity = str(upd.get("activity"))
                    pct = float(upd.get("percent"))
                    col_num = header_to_col.get(activity)
                    if not col_num:
                        continue

                    raw_value = from_percent(pct, phase)
                    row = rows.get(row_num)
                    if row is None:
                        continue

                    target_ref = f"{_xlsx_col_letter(col_num)}{row_num}"
                    cells = list(row.findall(f"{{{ns_main}}}c"))
                    cell = next((c for c in cells if c.attrib.get("r") == target_ref), None)

                    if cell is None:
                        cell = ET.Element(f"{{{ns_main}}}c", {"r": target_ref})
                        # Copiar estilo de una celda cercana de la misma fila.
                        nearest = None
                        nearest_dist = 10**9
                        for other in cells:
                            ref = other.attrib.get("r", "")
                            oc = _xlsx_col_number(ref)
                            if oc >= 5:
                                d = abs(oc - col_num)
                                if d < nearest_dist:
                                    nearest = other
                                    nearest_dist = d
                        if nearest is not None and "s" in nearest.attrib:
                            cell.set("s", nearest.attrib["s"])

                        inserted = False
                        for idx, other in enumerate(cells):
                            if _xlsx_col_number(other.attrib.get("r", "")) > col_num:
                                row.insert(idx, cell)
                                inserted = True
                                break
                        if not inserted:
                            row.append(cell)

                    # Convertir a número manteniendo el estilo.
                    cell.attrib.pop("t", None)
                    for child in list(cell):
                        if child.tag in {
                            f"{{{ns_main}}}f",
                            f"{{{ns_main}}}v",
                            f"{{{ns_main}}}is",
                        }:
                            cell.remove(child)
                    v = ET.SubElement(cell, f"{{{ns_main}}}v")
                    v.text = str(raw_value)

                except Exception:
                    continue

            replacements[path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        # Forzar recálculo de fórmulas al abrir Excel.
        calc_pr = workbook_root.find(f"{{{ns_main}}}calcPr")
        if calc_pr is None:
            calc_pr = ET.SubElement(workbook_root, f"{{{ns_main}}}calcPr")
        calc_pr.set("calcMode", "auto")
        calc_pr.set("fullCalcOnLoad", "1")
        calc_pr.set("forceFullCalc", "1")
        replacements["xl/workbook.xml"] = ET.tostring(
            workbook_root, encoding="utf-8", xml_declaration=True
        )

        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = replacements.get(item.filename)
                if data is None:
                    data = zin.read(item.filename)
                zout.writestr(item, data)

    output.seek(0)
    return output.getvalue()


def summary_display(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["RESUMEN"]
    data = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=data[0])

    pct_cols = [c for c in df.columns if c not in ["Fase","Partidas "]]
    def phase_name(v):
        s = str(v).strip().replace(" ","").upper()
        return s if s in PHASES else None

    for idx,row in df.iterrows():
        phase = phase_name(row.get("Fase"))
        if not phase:
            continue
        for col in pct_cols:
            try:
                v = row[col]
                if v is None or v == "":
                    df.at[idx,col] = ""
                else:
                    x=float(v)
                    if phase_scale(phase)==1.0:
                        x*=100
                    df.at[idx,col]=f"{x:.1f}%"
            except:
                pass
    return df

def normalize_route_pct(value):
    try:
        if value is None or value == "":
            return 0.0
        x = float(value)
        if 0 <= x <= 1:
            return x * 100.0
        if 0 <= x <= 100:
            return x
        return 0.0
    except Exception:
        return 0.0

@st.cache_data(show_spinner=False)
def route_critical_data(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    if "SEGUIMIENTO R.CRITICA" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb["SEGUIMIENTO R.CRITICA"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) <= 1:
        return pd.DataFrame()

    headers = [
        v if v is not None else f"Columna {i+1}"
        for i, v in enumerate(rows[0])
    ]
    df = pd.DataFrame(rows[1:], columns=headers)

    if "Piso" in df.columns:
        df = df[df["Piso"].notna()].copy()

    id_cols = ["Columna 1", "Piso", "Torre", "Departamento"]
    old_progress = "Av. Depto/Piso"
    activity_cols = [c for c in df.columns if c not in id_cols and c != old_progress]

    for col in activity_cols:
        df[col] = df[col].map(normalize_route_pct)

    if activity_cols:
        df["Avance Ruta Crítica"] = df[activity_cols].mean(axis=1).round(1)
    else:
        df["Avance Ruta Crítica"] = 0.0

    return df

def route_detail_only(path):
    df = route_critical_data(path).copy()
    if df.empty:
        return df
    if "Departamento" in df.columns:
        df = df[df["Departamento"].astype(str).str.strip().str.lower() != "todos"].copy()
    return df

def route_overall(path):
    df = route_detail_only(path)
    if df.empty or "Avance Ruta Crítica" not in df.columns:
        return 0.0
    vals = pd.to_numeric(df["Avance Ruta Crítica"], errors="coerce").dropna()
    return round(float(vals.mean()), 1) if len(vals) else 0.0

def route_by_floor(path):
    df = route_detail_only(path)
    if df.empty or "Piso" not in df.columns:
        return pd.DataFrame()
    out = df.groupby("Piso", dropna=True)["Avance Ruta Crítica"].mean().reset_index()
    out = out.rename(columns={"Avance Ruta Crítica": "Avance (%)"})
    out["Avance (%)"] = out["Avance (%)"].round(1)
    return out

def route_by_tower(path):
    df = route_detail_only(path)
    if df.empty or "Torre" not in df.columns:
        return pd.DataFrame()
    out = df.groupby("Torre", dropna=True)["Avance Ruta Crítica"].mean().reset_index()
    out = out.rename(columns={"Avance Ruta Crítica": "Avance (%)"})
    out["Avance (%)"] = out["Avance (%)"].round(1)
    return out

def route_by_activity(path):
    df = route_detail_only(path)
    if df.empty:
        return pd.DataFrame()
    skip = {"Columna 1","Piso","Torre","Departamento","Av. Depto/Piso","Avance Ruta Crítica"}
    rows = []
    for col in df.columns:
        if col in skip:
            continue
        vals = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(vals):
            rows.append({"Partida": col, "Avance (%)": round(float(vals.mean()), 1)})
    return pd.DataFrame(rows).sort_values("Avance (%)", ascending=False)

def route_display(path, torre="Todas", piso="Todos"):
    df = route_detail_only(path).copy()
    if df.empty:
        return df
    if torre != "Todas" and "Torre" in df.columns:
        df = df[df["Torre"].astype(str) == str(torre)]
    if piso != "Todos" and "Piso" in df.columns:
        df = df[df["Piso"] == piso]

    drop_cols = [c for c in ["Columna 1", "Av. Depto/Piso"] if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)

    id_cols = ["Piso","Torre","Departamento"]
    for col in df.columns:
        if col not in id_cols:
            vals = pd.to_numeric(df[col], errors="coerce")
            df[col] = vals.map(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
    return df

def ensure_history_sheet(path):
    wb = load_workbook(path)
    if "HISTORIAL_SEMANAL" not in wb.sheetnames:
        ws = wb.create_sheet("HISTORIAL_SEMANAL")
        ws.append([
            "Fecha actualización","Avance General","Fase 1","Fase 2","Fase 3","Fase 4","Usuario"
        ])
    wb.save(path)
    wb.close()

def register_weekly_snapshot(path, summaries, general, username, force=False):
    now = datetime.now()
    if now.weekday() != 4 and not force:
        return False, "La actualización semanal corresponde a los viernes."

    if db_ready():
        ok, msg = db_register_weekly(summaries, general, username)
        if not ok:
            return False, msg
    else:
        msg = "Base online no configurada; se guardará solo en el Excel de esta sesión."

    ensure_history_sheet(path)
    date_value = now.strftime("%Y-%m-%d")
    wb = load_workbook(path)
    ws = wb["HISTORIAL_SEMANAL"]
    target_row = None
    for r in range(2, ws.max_row + 1):
        existing = ws.cell(r, 1).value
        if existing is not None and str(existing)[:10] == date_value:
            target_row = r
            break
    vals = [
        date_value,float(general),float(summaries["FASE1"]),float(summaries["FASE2"]),
        float(summaries["FASE3"]),float(summaries["FASE4"]),username
    ]
    if target_row:
        for c,v in enumerate(vals, start=1):
            ws.cell(target_row,c).value = v
    else:
        ws.append(vals)
    wb.save(path)
    wb.close()
    return True, msg

@st.cache_data(show_spinner=False, ttl=5)
def load_weekly_history(path):
    rows = db_weekly_history()
    if rows:
        df = pd.DataFrame(rows).rename(columns={
            "update_date":"Fecha actualización",
            "general":"Avance General",
            "fase1":"Fase 1",
            "fase2":"Fase 2",
            "fase3":"Fase 3",
            "fase4":"Fase 4",
            "updated_by":"Usuario",
        })
        df["Fecha actualización"] = pd.to_datetime(df["Fecha actualización"], errors="coerce")
        return df.sort_values("Fecha actualización")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "HISTORIAL_SEMANAL" not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()
        ws = wb["HISTORIAL_SEMANAL"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) <= 1:
            return pd.DataFrame()
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df["Fecha actualización"] = pd.to_datetime(df["Fecha actualización"], errors="coerce")
        return df.sort_values("Fecha actualización")
    except Exception:
        return pd.DataFrame()

def next_friday_text():
    today = datetime.now().date()
    days = (4 - today.weekday()) % 7
    if days == 0:
        return "Hoy corresponde actualización semanal."
    nxt = today + pd.Timedelta(days=days)
    return f"Próxima actualización: viernes {nxt.strftime('%d-%m-%Y')}."

def raw_sheet_df(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name]
    data = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    width=max((len(r) for r in data), default=1)
    data=[r+[None]*(width-len(r)) for r in data]
    return pd.DataFrame(data, columns=[f"Columna {i+1}" for i in range(width)])


def current_role():
    return st.session_state.get("user_role", "Usuario")

def current_phase():
    return str(st.session_state.get("user_phase", "ALL")).upper()

def allowed_phases():
    role = current_role()
    phase = current_phase()
    if role in ["Administrador", "Visor"]:
        return PHASES
    if role == "Editor" and phase in PHASES:
        return [phase]
    return []

def can_edit_phase(phase):
    return current_role() == "Administrador" or (
        current_role() == "Editor" and current_phase() == phase
    )

def can_access_admin_tools():
    return current_role() == "Administrador"

with st.sidebar:
    if os.path.exists(LOGO):
        st.image(LOGO, use_container_width=True)

    st.markdown("## CONTROL DE OBRA")
    st.caption("EDIFICIO SAN FRANCISCO 211 · PAZ")
    st.markdown("---")

    role = current_role()
    assigned_phase = current_phase()

    if role == "Administrador":
        menu_options = [
            "📊 Dashboard",
            "📆 Comparación semanal",
            "📈 Gráficos por fase",
            "📅 Carta Gantt",
            "🧱 Actualizar avances",
            "🏢 Fases completas",
            "🏙️ Avance por Departamento",
            "⚠️ Ruta crítica",
            "⬆️ Importar / Exportar",
        ]
    elif role == "Editor":
        menu_options = [
            "📊 Dashboard",
            "📈 Gráficos por fase",
            "📅 Carta Gantt",
            "🧱 Actualizar avances",
            "🏢 Fases completas",
            "🏙️ Avance por Departamento",
        ]
    elif role == "Visor":
        menu_options = [
            "📊 Dashboard",
            "📆 Comparación semanal",
            "📈 Gráficos por fase",
            "📅 Carta Gantt",
            "🏢 Fases completas",
            "🏙️ Avance por Departamento",
            "⚠️ Ruta crítica",
        ]
    else:
        menu_options = ["📊 Dashboard"]

    page = st.radio("Menú", menu_options, label_visibility="collapsed")

    st.markdown("---")
    st.caption(f"👤 {st.session_state.get('user_name','Usuario')}")
    st.caption(f"Rol: {role}")
    st.caption("Acceso: Todas las fases" if assigned_phase == "ALL"
               else f"Acceso: {assigned_phase.replace('FASE','Fase ')}")

    ok_db, db_msg = supabase_status()
    st.caption("🟢 Base online conectada" if ok_db else "🟠 Base online con error")

    if st.button("🚪 Cerrar sesión", use_container_width=True):
        for k in ["authenticated","username","user_name","user_role","user_phase"]:
            st.session_state.pop(k, None)
        st.rerun()

hlogo, htitle = st.columns([1.1,4.9], vertical_alignment="center")
with hlogo:
    if os.path.exists(LOGO):
        st.image(LOGO, use_container_width=True)
with htitle:
    st.markdown('<div class="sf-title">CONTROL FASES SAN FRANCISCO 211</div>', unsafe_allow_html=True)
st.markdown(
    f'<div class="sf-sub">Panel de control profesional · v41 recuperación auditada · {datetime.now().strftime("%d-%m-%Y %H:%M")}</div>',
    unsafe_allow_html=True,
)

sync_a, sync_b = st.columns([4,1])
with sync_b:
    if st.button("🔄 Sincronizar datos", use_container_width=True, key="global_sync"):
        st.cache_data.clear()
        st.rerun()
with sync_a:
    online_now = db_phase_updates()
    if online_now:
        stamps = [str(x.get("updated_at","")) for x in online_now if x.get("updated_at")]
        last_online = max(stamps) if stamps else ""
        st.caption(f"🟢 Base online sincronizada · última modificación: {last_online[:19].replace('T',' ')}")
    else:
        st.caption("🟢 Base online conectada · sin cambios detallados registrados")

# Avance calculado desde el detalle actual (partidas/departamentos).
live_summaries = {p: phase_summary(st.session_state.workbook_path, p) for p in PHASES}
live_general = round(sum(live_summaries.values()) / 4, 1)

# FUENTES DE AVANCE:
# - live_summaries: cálculo crudo desde el detalle disponible en Excel + phase_updates.
# - weekly_history: último corte histórico confirmado.
# PROTECCIÓN DE CONTINUIDAD: una actualización de versión o un registro faltante en
# phase_updates no puede hacer retroceder el Dashboard por debajo del último corte
# confirmado. El corte semanal se usa como piso de recuperación, no como sustituto
# del avance vivo.
summaries = dict(live_summaries)
general = live_general
official_snapshot_date = None
official_snapshot_user = None
official_snapshot_source = "weekly_history · Supabase"
official_summaries = None
official_general = None
_latest_official = db_latest_weekly_snapshot()
if _latest_official:
    try:
        official_summaries = {
            "FASE1": float(pd.to_numeric(_latest_official.get("fase1"), errors="coerce") or 0),
            "FASE2": float(pd.to_numeric(_latest_official.get("fase2"), errors="coerce") or 0),
            "FASE3": float(pd.to_numeric(_latest_official.get("fase3"), errors="coerce") or 0),
            "FASE4": float(pd.to_numeric(_latest_official.get("fase4"), errors="coerce") or 0),
        }
        official_general = float(pd.to_numeric(_latest_official.get("general"), errors="coerce") or 0)
        official_snapshot_date = pd.to_datetime(_latest_official.get("update_date"), errors="coerce")
        official_snapshot_user = _latest_official.get("updated_by")
    except Exception:
        official_summaries = None
        official_general = None

# Recuperación segura del avance: si el detalle online está incompleto, nunca mostrar
# menos que el último corte semanal confirmado. Esto evita regresiones como 19→17 o 8→4
# causadas por registros parciales de phase_updates.
if official_summaries is not None:
    summaries = {
        p: round(max(float(live_summaries.get(p, 0.0)), float(official_summaries.get(p, 0.0))), 1)
        for p in PHASES
    }
    general = round(max(sum(summaries.values()) / 4.0, float(official_general or 0.0)), 1)
else:
    summaries = dict(live_summaries)
    general = live_general

# AVANCE GENERAL SIEMPRE VISIBLE EN TODAS LAS PÁGINAS
st.markdown('<div class="section">AVANCE GENERAL DEL PROYECTO</div>', unsafe_allow_html=True)
gcol, p1col, p2col, p3col, p4col = st.columns([1.15,1,1,1,1])

with gcol:
    st.markdown(
        f'<div class="general-box"><div class="general-label">Avance General</div>'
        f'<div class="general-pct">{general:.1f}%</div></div>',
        unsafe_allow_html=True
    )
    st.progress(min(max(general/100,0),1))

for col, phase in zip([p1col,p2col,p3col,p4col], PHASES):
    val = summaries[phase]
    with col:
        st.markdown(
            f'<div class="phase-card"><div class="phase-name">{phase.replace("FASE","Fase ")}</div>'
            f'<div class="phase-pct">{val:.1f}%</div></div>',
            unsafe_allow_html=True
        )
        st.progress(min(max(val/100,0),1))

st.caption(
    f"🟢 Avance actual en vivo · General {general:.1f}% · "
    f"F1 {summaries['FASE1']:.1f}% · F2 {summaries['FASE2']:.1f}% · "
    f"F3 {summaries['FASE3']:.1f}% · F4 {summaries['FASE4']:.1f}% · "
    "Fuente: cambios guardados en Supabase + protección del último corte confirmado."
)
if official_summaries is not None and any(float(live_summaries[p]) < float(official_summaries[p]) for p in PHASES):
    st.caption("🛡️ Recuperación activa: el detalle online está incompleto en una o más fases; se conserva como mínimo el último avance semanal confirmado para evitar pérdida visual de avance.")
if official_snapshot_date is not None and not pd.isna(official_snapshot_date) and official_summaries is not None:
    _who = f" · registrado por {official_snapshot_user}" if official_snapshot_user else ""
    st.caption(
        f"📅 Último corte semanal: {official_snapshot_date.strftime('%d-%m-%Y')}{_who} · "
        f"General {official_general:.1f}% · F1 {official_summaries['FASE1']:.1f}% · "
        f"F2 {official_summaries['FASE2']:.1f}% · F3 {official_summaries['FASE3']:.1f}% · "
        f"F4 {official_summaries['FASE4']:.1f}%."
    )
st.divider()



def _phase_hex(ph):
    return {"FASE1":"#0B63B6","FASE2":"#11879A","FASE3":"#2E8B2E","FASE4":"#E3A500"}.get(ph,"#64748B")

def _status_label(pct):
    try: x=float(pct)
    except Exception: x=0.0
    if x >= 99.999: return "✅ Al día"
    if x >= 50: return "⚠️ En progreso"
    if x > 0: return "⚠️ En riesgo"
    return "🕘 No iniciado"

def route_activity_summary(path, phases):
    frames=[]
    for ph in phases:
        m=critical_route_matrix(path, ph)
        if m.empty: continue
        x=m[["Fase","Partida","% Avance Real"]].copy()
        x["Estado"] = x["% Avance Real"].map(_status_label)
        x["Crítica"] = "🔴 Sí"
        frames.append(x)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

def render_route_activity_html(df):
    if df.empty:
        return '<div class="route-empty">Sin datos de ruta crítica.</div>'
    phase_colors={"FASE 1":"#0B63B6","FASE 2":"#11879A","FASE 3":"#2E8B2E","FASE 4":"#E3A500"}
    rows=[]; last=None
    for _,r in df.iterrows():
        ph=str(r["Fase"]); pct=float(r["% Avance Real"])
        bar = '#22c55e' if pct >= 99.999 else ('#84cc16' if pct >= 50 else ('#fbbf24' if pct > 0 else '#cbd5e1'))
        phase_cell = ('<td class="phase-band" style="background:%s">%s</td>' % (phase_colors.get(ph,'#64748B'), ph)) if ph!=last else '<td class="phase-band phase-blank"></td>'
        last=ph
        rows.append('<tr>%s<td class="task-name">%s</td><td class="pct-cell"><div class="mini-track"><div class="mini-fill" style="width:%.1f%%;background:%s"></div></div><span>%.0f%%</span></td><td class="state-cell">%s</td><td class="critical-cell">%s</td></tr>' % (phase_cell, str(r["Partida"]), max(0,min(100,pct)), bar, pct, str(r["Estado"]), str(r["Crítica"])))
    return '<div class="route-table-wrap"><table class="route-table"><thead><tr><th>FASE</th><th>PARTIDA (ACTIVIDAD)</th><th>% AVANCE REAL</th><th>ESTADO</th><th>CRÍTICA</th></tr></thead><tbody>'+''.join(rows)+'</tbody></table></div>'

if page == "📊 Dashboard":
    ok_db, db_msg = supabase_status()
    if ok_db:
        st.success("🟢 Supabase conectado correctamente.")
    else:
        st.warning("🟠 Supabase aún no está conectado.")
        with st.expander("Diagnóstico de conexión"):
            st.code(db_msg)
            st.markdown(
                "Formato esperado en Streamlit Secrets:\n\n"
                "[supabase]\n"
                'url = "https://TU-PROYECTO.supabase.co"\n'
                'service_role_key = "sb_secret_..."'
            )

    st.success(
        f"🔄 Dashboard en vivo · General {general:.1f}% · "
        f"Fase 1 {summaries['FASE1']:.1f}% · Fase 2 {summaries['FASE2']:.1f}% · "
        f"Fase 3 {summaries['FASE3']:.1f}% · Fase 4 {summaries['FASE4']:.1f}%. "
        "Se recalcula desde las partidas guardadas en Supabase."
    )
    if official_snapshot_date is not None and not pd.isna(official_snapshot_date) and official_summaries is not None:
        st.caption(
            f"Referencia histórica · último corte semanal {official_snapshot_date.strftime('%d-%m-%Y')}: "
            f"General {official_general:.1f}% · F1 {official_summaries['FASE1']:.1f}% · "
            f"F2 {official_summaries['FASE2']:.1f}% · F3 {official_summaries['FASE3']:.1f}% · "
            f"F4 {official_summaries['FASE4']:.1f}%."
        )
    c1,c2,c3,c4,c5=st.columns(5)
    cards=[
        (c1,"AVANCE GENERAL",f"{general:.1f}%","blue"),
        (c2,"FASE 1",f"{summaries['FASE1']:.0f}%","good"),
        (c3,"FASE 2",f"{summaries['FASE2']:.0f}%","warn"),
        (c4,"FASE 3",f"{summaries['FASE3']:.0f}%","warn"),
        (c5,"FASE 4",f"{summaries['FASE4']:.0f}%","bad"),
    ]
    for c,lab,val,cl in cards:
        c.markdown(f'<div class="card"><div class="klabel">{lab}</div><div class="kvalue {cl}">{val}</div></div>',unsafe_allow_html=True)

    st.write("")
    left,right=st.columns([1.25,1])
    with left:
        st.markdown('<div class="section">AVANCE GENERAL POR FASE · DESDE FASES COMPLETAS</div>',unsafe_allow_html=True)
        chart=pd.DataFrame({
            "Fase":["Fase 1","Fase 2","Fase 3","Fase 4"],
            "Avance (%)":[summaries[p] for p in PHASES],
        })
        st.bar_chart(chart.set_index("Fase"),height=330)
        compare = chart.copy()
        compare["Avance"] = compare["Avance (%)"].map(lambda x: f"{x:.1f}%")
        compare_num = chart[["Fase","Avance (%)"]].copy()
        st.dataframe(style_percent_df(compare_num, ["Avance (%)"]), use_container_width=True, hide_index=True)
    with right:
        st.markdown('<div class="section">ESTADO ACTUAL</div>',unsafe_allow_html=True)
        for p in PHASES:
            v=summaries[p]
            st.write(f"**{p.replace('FASE','Fase ')} — {v:.1f}%**")
            st.progress(min(max(v/100,0),1))

    st.markdown('<div class="section">AVANCE POR PISO · DESDE FASES COMPLETAS</div>',unsafe_allow_html=True)
    selected_phase=st.selectbox("Fase para detalle",allowed_phases(),format_func=lambda x:x.replace("FASE","Fase "))
    piso=phase_by_floor(st.session_state.workbook_path,selected_phase)
    if not piso.empty:
        st.bar_chart(piso.set_index("Piso"),height=330)

elif page == "📆 Comparación semanal":
    st.markdown('<div class="section">COMPARACIÓN DE AVANCE · REGISTRO DE LOS VIERNES</div>', unsafe_allow_html=True)

    st.info(
        "Cada punto del gráfico corresponde al avance registrado en un viernes. "
        "Se compara el Avance General y las cuatro fases a través del tiempo."
    )

    is_friday = datetime.now().weekday() == 4
    role = st.session_state.get("user_role", "Usuario")
    user = st.session_state.get(
        "user_name",
        st.session_state.get("username", "Usuario")
    )

    top1, top2 = st.columns([2,1])

    with top1:
        if is_friday:
            st.success("✅ Hoy es viernes: corresponde registrar el avance semanal.")
        else:
            st.info("📅 " + next_friday_text())

    with top2:
        if current_role() == "Administrador" and st.button(
            "📌 REGISTRAR AVANCE DE HOY",
            type="primary",
            use_container_width=True
        ):
            ok, msg = register_weekly_snapshot(
                st.session_state.workbook_path,
                live_summaries,
                live_general,
                user,
                force=(role == "Administrador")
            )
            load_weekly_history.clear()

            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.warning(msg)

    if not is_friday and role == "Administrador":
        st.caption(
            "Administrador: puedes registrar manualmente fuera del viernes "
            "para regularizar o probar el historial."
        )

    # Mostrar siempre el avance ACTUAL EN VIVO, aunque el historial semanal no se haya registrado todavía.
    st.markdown("### Avance actual en vivo")
    a1,a2,a3,a4,a5 = st.columns(5)
    a1.metric("Avance General", f"{general:.1f}%")
    a2.metric("Fase 1", f"{summaries['FASE1']:.0f}%")
    a3.metric("Fase 2", f"{summaries['FASE2']:.0f}%")
    a4.metric("Fase 3", f"{summaries['FASE3']:.0f}%")
    a5.metric("Fase 4", f"{summaries['FASE4']:.0f}%")
    st.caption("Estos valores cambian inmediatamente al guardar avances. El bloque siguiente corresponde al historial registrado de los viernes.")

    hist = load_weekly_history(st.session_state.workbook_path)

    if hist.empty:
        st.warning(
            "Todavía no existen viernes registrados. "
            "Presiona «Registrar avance de hoy» para crear el primer punto del historial."
        )
    else:
        hist = hist.copy()
        hist["Fecha actualización"] = pd.to_datetime(
            hist["Fecha actualización"],
            errors="coerce"
        )
        hist = hist.dropna(subset=["Fecha actualización"]).sort_values("Fecha actualización")

        metric_cols = ["Avance General","Fase 1","Fase 2","Fase 3","Fase 4"]
        for c in metric_cols:
            hist[c] = pd.to_numeric(hist[c], errors="coerce")

        latest = hist.iloc[-1]

        st.markdown("### Último viernes registrado · histórico")
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("Avance General", f"{latest['Avance General']:.1f}%")
        k2.metric("Fase 1", f"{latest['Fase 1']:.0f}%")
        k3.metric("Fase 2", f"{latest['Fase 2']:.0f}%")
        k4.metric("Fase 3", f"{latest['Fase 3']:.0f}%")
        k5.metric("Fase 4", f"{latest['Fase 4']:.0f}%")

        chart_df = hist[["Fecha actualización"] + metric_cols].copy()
        chart_df["Viernes"] = chart_df["Fecha actualización"].dt.strftime("%d-%m-%Y")
        chart_df = chart_df.set_index("Viernes")[metric_cols]

        st.markdown("### Evolución semanal del proyecto")
        st.line_chart(chart_df, height=400)

        st.caption(
            "Eje horizontal: fecha de actualización de cada viernes. "
            "Eje vertical: porcentaje de avance 0%–100%."
        )

        if len(hist) >= 2:
            current = hist.iloc[-1]
            previous = hist.iloc[-2]

            st.markdown("### Comparación con el viernes anterior")

            comparison = pd.DataFrame({
                "Indicador": metric_cols,
                "Viernes anterior": [previous[c] for c in metric_cols],
                "Último viernes": [current[c] for c in metric_cols],
            })
            comparison["Variación (pp)"] = (
                comparison["Último viernes"] -
                comparison["Viernes anterior"]
            ).round(1)

            # Formato explícito en porcentaje para evitar que Streamlit muestre decimales crudos.
            comp_styler = comparison.style
            comp_styler = comp_styler.map(_pct_style, subset=["Viernes anterior","Último viernes"])
            comp_styler = comp_styler.format({
                "Viernes anterior": "{:.1f}%",
                "Último viernes": "{:.1f}%",
                "Variación (pp)": "{:+.1f} pp",
            }, na_rep="")
            st.dataframe(comp_styler, use_container_width=True, hide_index=True)

        st.markdown("### Historial de viernes registrados")

        shown = hist.copy()
        shown["Fecha actualización"] = shown["Fecha actualización"].dt.strftime("%d-%m-%Y")

        for c in metric_cols:
            shown[c] = shown[c].map(
                lambda x: f"{x:.1f}%" if pd.notna(x) else ""
            )

        visible_cols = [
            "Fecha actualización",
            "Avance General",
            "Fase 1",
            "Fase 2",
            "Fase 3",
            "Fase 4",
            "Usuario"
        ]
        visible_cols = [c for c in visible_cols if c in shown.columns]

        st.dataframe(
            shown[visible_cols],
            use_container_width=True,
            hide_index=True
        )

        st.success(
            f"Historial disponible: {len(hist)} actualización(es) registrada(s)."
        )

elif page == "📈 Gráficos por fase":
    st.markdown('<div class="section">GRÁFICOS DE CADA FASE · FUENTE: FASES COMPLETAS</div>',unsafe_allow_html=True)
    st.caption("Todos los gráficos se generan directamente desde los datos de «Fases completas» y trabajan en escala 0%–100%.")

    for phase in allowed_phases():
        st.markdown(f"## {phase.replace('FASE','Fase ')}")
        st.metric("AVANCE GENERAL DE LA FASE", f"{summaries[phase]:.0f}%")
        st.progress(min(max(summaries[phase]/100,0),1))
        a,b=st.columns(2)

        with a:
            st.markdown("**Avance por piso**")
            pf=phase_by_floor(st.session_state.workbook_path,phase)
            if not pf.empty:
                st.bar_chart(pf.set_index("Piso"),height=300)

        with b:
            st.markdown("**Avance por torre**")
            tw=phase_by_tower(st.session_state.workbook_path,phase)
            if not tw.empty:
                st.bar_chart(tw.set_index("Torre"),height=300)

        st.markdown("**Avance promedio por partida**")
        act=phase_activity_summary(st.session_state.workbook_path,phase)
        if not act.empty:
            st.bar_chart(act.set_index("Partida"),height=400)
        st.divider()

elif page == "🧱 Actualizar avances":
    if current_role() == "Visor":
        st.error("Tu perfil es solo de lectura.")
        st.stop()
    st.markdown('<div class="section">ACTUALIZAR AVANCE DE DEPARTAMENTO</div>',unsafe_allow_html=True)
    st.success("Todos los controles de avance se muestran y editan de 0% a 100%.")

    phase=st.selectbox("Fase",allowed_phases(),format_func=lambda x:x.replace("FASE","Fase "))
    df=load_phase(st.session_state.workbook_path,phase)

    towers=sorted([str(x) for x in df["Torre"].dropna().unique()])
    torre=st.selectbox("Torre",towers)
    f1=df[df["Torre"].astype(str)==torre]

    pisos=sorted(f1["Piso"].dropna().unique(),key=lambda x:float(x) if str(x).replace(".","",1).isdigit() else str(x))
    piso=st.selectbox("Piso",pisos)
    f2=f1[f1["Piso"]==piso]

    deps=sorted(f2["Departamento"].dropna().unique(),key=lambda x:str(x))
    depto=st.selectbox("Departamento",deps)
    rec=f2[f2["Departamento"]==depto].iloc[0]
    excel_row=int(rec["_excel_row"])

    progress_col="% Avance Real Depto"
    current=to_percent(rec[progress_col],phase)
    st.metric("Avance actual del departamento",f"{current:.1f}%")
    st.progress(min(max(current/100,0),1))

    skip=["Fase","Piso","Torre","Departamento",progress_col,"_excel_row"]
    activities=[c for c in df.columns if c not in skip]

    st.markdown('<div class="section">PARTIDAS</div>',unsafe_allow_html=True)
    edits={}
    cols=st.columns(2)
    for i,activity in enumerate(activities):
        val=to_percent(rec[activity],phase)
        with cols[i%2]:
            edits[activity]=st.number_input(
                f"{activity} (%)",
                min_value=0.0,max_value=100.0,
                value=float(round(val,1)),step=5.0,format="%.1f",
                key=f"{phase}-{excel_row}-{i}"
            )

    calculated=round(sum(edits.values())/len(edits),1) if edits else 0.0
    st.markdown(f"### Nuevo avance calculado: **{calculated:.1f}%**")
    st.progress(min(max(calculated/100,0),1))

    st.caption("☁️ Autoguardado online activo: cada cambio se guarda y verifica automáticamente en Supabase.")
    if not can_edit_phase(phase):
        st.warning("No tienes permiso para modificar esta fase.")
    else:
        changed = {}
        for activity in activities:
            original_pct = round(to_percent(rec[activity], phase), 1)
            new_pct = round(float(edits[activity]), 1)
            if abs(new_pct - original_pct) >= 0.05:
                changed[activity] = new_pct
        if changed:
            with st.spinner("Guardando cambio en línea…"):
                saved, errors = save_department(st.session_state.workbook_path, phase, excel_row, changed)
            if errors:
                show_save_status()
                st.stop()
            st.session_state["autosave_notice"] = f"{saved} cambio(s) guardado(s) en Depto {depto}."
            st.rerun()
        if st.session_state.pop("autosave_notice", None):
            show_save_status()
        else:
            show_save_status()

elif page == "🏢 Fases completas":
    st.markdown('<div class="section">EDITAR FASE COMPLETA</div>', unsafe_allow_html=True)
    st.success("Ahora puedes editar directamente todas las partidas de cada Fase. Los valores se trabajan de 0% a 100%.")

    phase = st.selectbox(
        "Selecciona una fase",
        allowed_phases(),
        format_func=lambda x:x.replace("FASE","Fase ")
    )

    original = phase_editor_df(st.session_state.workbook_path, phase)

    # Indicadores calculados desde la misma fuente de datos que alimenta la tabla.
    current_phase_pct = phase_summary(st.session_state.workbook_path, phase)
    cphase, cdept = st.columns(2)
    cphase.metric("Avance actual de la fase", f"{current_phase_pct:.1f}%")
    if "% Avance Real Depto" in original.columns:
        dept_vals = pd.to_numeric(original["% Avance Real Depto"], errors="coerce").dropna()
        cdept.metric("Promedio departamentos", f"{float(dept_vals.mean()) if len(dept_vals) else 0.0:.1f}%")

    # No mostrar _excel_row en pantalla.
    visible_cols = [c for c in original.columns if c != "_excel_row"]
    editor_source = original[visible_cols].copy()

    # Proteger columnas identificadoras y el cálculo total.
    disabled_cols = ["Fase", "Piso", "Torre", "Departamento", "% Avance Real Depto"]

    st.caption(
        "Puedes modificar las partidas. Las columnas Fase, Piso, Torre, Departamento y "
        "% Avance Real Depto quedan protegidas; el avance total se recalcula automáticamente después de cada autoguardado."
    )

    column_cfg = {}
    for c in visible_cols:
        if c not in ["Fase", "Piso", "Torre", "Departamento"]:
            column_cfg[c] = st.column_config.NumberColumn(
                c,
                min_value=0.0,
                max_value=100.0,
                step=5.0,
                format="%.1f%%"
            )

    # Cambiar la identidad del editor después de cada autoguardado evita que
    # Streamlit conserve en memoria el antiguo % Avance Real Depto.
    revision_key = f"full_editor_revision_{phase}"
    if revision_key not in st.session_state:
        st.session_state[revision_key] = 0
    editor_key = f"full_editor_{phase}_{st.session_state[revision_key]}"
    edited = st.data_editor(
        editor_source,
        use_container_width=True,
        height=700,
        num_rows="fixed",
        disabled=disabled_cols,
        column_config=column_cfg,
        key=editor_key
    )

    st.caption("☁️ Autoguardado online activo: al modificar una celda, el cambio se guarda y verifica en Supabase automáticamente.")
    if can_edit_phase(phase):
        editable_cols = [c for c in visible_cols if c not in disabled_cols]
        has_changes = False
        for c in editable_cols:
            a = pd.to_numeric(edited[c], errors="coerce").fillna(0.0)
            b = pd.to_numeric(editor_source[c], errors="coerce").fillna(0.0)
            if ((a - b).abs() >= 0.05).any():
                has_changes = True
                break
        if has_changes:
            edited_full = original.copy()
            for c in visible_cols:
                edited_full[c] = edited[c]
            try:
                with st.spinner("Autoguardando cambios de la fase…"):
                    changed_cells, changed_rows = save_full_phase(
                        st.session_state.workbook_path, phase, edited_full, original
                    )
                if changed_cells:
                    # El guardado ya fue confirmado en Supabase. Se invalida el editor
                    # para que vuelva a leer la base online y recalcule inmediatamente
                    # % Avance Real Depto, piso y fase.
                    st.session_state[revision_key] += 1
                    st.session_state["phase_autosave_notice"] = (
                        f"{changed_cells} celda(s) · {changed_rows} departamento(s) · porcentajes recalculados"
                    )
                    st.rerun()
            except Exception as e:
                set_save_status(False, str(e), phase)
                show_save_status()
    if st.session_state.pop("phase_autosave_notice", None):
        show_save_status()
    else:
        show_save_status()

    with st.expander("🎨 Vista semáforo de porcentajes", expanded=False):
        preview_cols = [c for c in visible_cols if c not in ["Fase","Piso","Torre","Departamento"]]
        st.caption("Verde = 100% · Amarillo = 50% a 99% · Naranjo = 0% a 49%")
        st.dataframe(
            style_percent_df(editor_source[visible_cols], preview_cols),
            use_container_width=True, height=520, hide_index=True
        )

    csave, cinfo = st.columns([1,2])
    with csave:
        if not can_edit_phase(phase):
            st.info("Modo solo lectura: no tienes permiso para modificar esta fase.")
        elif st.button("🔎 VERIFICAR CONEXIÓN Y ÚLTIMO GUARDADO", use_container_width=True):
            ok_db, db_msg = supabase_status()
            if ok_db:
                rows = [r for r in db_phase_updates() if r.get("phase") == phase]
                if rows:
                    latest = max(rows, key=lambda r: str(r.get("updated_at", "")))
                    st.success(
                        f"Base online conectada. Último cambio: {latest.get('updated_at','—')} · "
                        f"{latest.get('updated_by','—')} · {latest.get('activity','—')} = {float(latest.get('percent',0)):.1f}%"
                    )
                else:
                    st.info("Base online conectada, sin cambios guardados todavía para esta fase.")
            else:
                st.error(f"No hay conexión con Supabase: {db_msg}")

    with cinfo:
        st.info(
            "No es necesario presionar Guardar. Cada celda modificada se autoguarda en la base online, "
            "se verifica y luego se recalculan Dashboard, Ruta Crítica y Gantt."
        )

elif page == "🏙️ Avance por Departamento":
    st.markdown('<div class="section">AVANCE POR DEPARTAMENTO · ELEVACIÓN DEL EDIFICIO</div>', unsafe_allow_html=True)
    st.caption(
        "Vista por Torre y Piso basada en la elevación del proyecto. Debajo de cada departamento se muestra "
        "un único promedio general por departamento, calculado como el promedio de Fase 1, Fase 2, Fase 3 y Fase 4 desde las mismas partidas guardadas en Supabase."
    )

    # Construir una única base por Torre + Piso + Departamento con el promedio real de cada fase.
    phase_frames = []
    for _phase in ["FASE1", "FASE2", "FASE3", "FASE4"]:
        _df = phase_numeric_percent_df(st.session_state.workbook_path, _phase).copy()
        if _df.empty:
            continue
        _df["Piso"] = pd.to_numeric(_df["Piso"], errors="coerce")
        _df["Departamento"] = pd.to_numeric(_df["Departamento"], errors="coerce")
        _df = _df.dropna(subset=["Piso", "Torre", "Departamento"])
        _df["Piso"] = _df["Piso"].astype(int)
        _df["Departamento"] = _df["Departamento"].astype(int)
        # Si la planilla trae filas repetidas, consolidar el departamento antes de dibujarlo.
        _g = (_df.groupby(["Torre", "Piso", "Departamento"], as_index=False)["% Avance Real Depto"]
              .mean()
              .rename(columns={"% Avance Real Depto": _phase}))
        phase_frames.append(_g)

    if not phase_frames:
        st.warning("No hay departamentos disponibles para mostrar.")
    else:
        from functools import reduce
        _all = reduce(
            lambda left, right: pd.merge(left, right, on=["Torre","Piso","Departamento"], how="outer"),
            phase_frames
        )
        for _phase in ["FASE1","FASE2","FASE3","FASE4"]:
            if _phase not in _all.columns:
                _all[_phase] = 0.0
            _all[_phase] = pd.to_numeric(_all[_phase], errors="coerce").fillna(0.0).clip(0,100)
        _all["Torre"] = _all["Torre"].astype(str).str.strip().str.upper()

        def _pct_style(v):
            v = float(v or 0)
            if v >= 99.95:
                return "#16a34a", "#ffffff"
            if v >= 50:
                return "#facc15", "#111827"
            return "#f97316", "#ffffff"

        st.markdown("""
        <style>
        .elev-tower {background:#f4f7fb;border:1px solid #dbe3ee;border-radius:12px;padding:12px 12px 16px;margin:8px 0 22px;}
        .elev-title {font-size:22px;font-weight:800;color:#17324d;margin:0 0 10px 2px;}
        .elev-floor {display:grid;grid-template-columns:72px 1fr;gap:8px;align-items:stretch;margin:7px 0;}
        .elev-floor-label {background:#263746;color:white;border-radius:7px;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:13px;min-height:88px;}
        .elev-depts {display:grid;grid-template-columns:repeat(auto-fit,minmax(118px,1fr));gap:6px;}
        .elev-dept {background:white;border:1px solid #cfd8e3;border-radius:7px;padding:7px;box-shadow:0 1px 2px rgba(0,0,0,.05);min-width:0;}
        .elev-dept-name {font-size:13px;font-weight:800;text-align:center;color:#1f2937;border-bottom:1px solid #e5e7eb;padding-bottom:5px;margin-bottom:5px;}
        .elev-phases {display:block;}
        .elev-phase {border-radius:6px;padding:7px 4px;text-align:center;font-size:12px;font-weight:800;white-space:nowrap;}
        @media (max-width: 900px){.elev-depts{grid-template-columns:repeat(3,minmax(105px,1fr));}}
        </style>
        """, unsafe_allow_html=True)

        _towers = sorted([t for t in _all["Torre"].dropna().unique().tolist() if t and t != "NAN"])
        _choice = st.selectbox("Torre", ["Todas"] + _towers, key="elev_tower_filter")
        _show_towers = _towers if _choice == "Todas" else [_choice]

        # Resumen superior por torre usando el mismo universo de departamentos visible.
        _summary_cols = st.columns(max(1, min(4, len(_show_towers))))
        for _i, _tower in enumerate(_show_towers):
            _td = _all[_all["Torre"] == _tower]
            _overall = float(_td[["FASE1","FASE2","FASE3","FASE4"]].mean(axis=1).mean()) if len(_td) else 0.0
            _summary_cols[_i % len(_summary_cols)].metric(f"Torre {_tower}", f"{_overall:.1f}%", f"{len(_td)} deptos")

        for _tower in _show_towers:
            _td = _all[_all["Torre"] == _tower].copy()
            html = [f'<div class="elev-tower"><div class="elev-title">TORRE {_tower}</div>']
            for _floor in sorted(_td["Piso"].dropna().astype(int).unique().tolist(), reverse=True):
                _fd = _td[_td["Piso"] == _floor].sort_values("Departamento")
                html.append(f'<div class="elev-floor"><div class="elev-floor-label">PISO {_floor}</div><div class="elev-depts">')
                for _, _r in _fd.iterrows():
                    html.append(f'<div class="elev-dept"><div class="elev-dept-name">DEPTO {int(_r["Departamento"])}</div><div class="elev-phases">')
                    _v = float(pd.Series([_r["FASE1"], _r["FASE2"], _r["FASE3"], _r["FASE4"]]).mean())
                    _bg, _fg = _pct_style(_v)
                    html.append(f'<div class="elev-phase" style="background:{_bg};color:{_fg}">{_v:.1f}%</div>')
                    html.append('</div></div>')
                html.append('</div></div>')
            html.append('</div>')
            st.markdown(''.join(html), unsafe_allow_html=True)

        st.caption("Cada departamento muestra solamente su promedio final. Semáforo: 🟢 100% · 🟡 50%–99,9% · 🟠 0%–49,9%. Se recalcula desde Supabase al actualizar la página.")

elif page == "📋 Resumen":
    st.caption("El RESUMEN se convierte automáticamente a porcentaje según la fase.")
    st.dataframe(summary_display(st.session_state.workbook_path),use_container_width=True,height=700)

elif page == "📅 Avance semanal":
    names=get_sheet_names(st.session_state.workbook_path)
    opts=[n for n in names if "AVANCE SEMANAL" in n.upper()]
    sh=opts[0] if opts else names[0]
    st.dataframe(raw_sheet_df(st.session_state.workbook_path,sh),use_container_width=True,height=700)

elif page == "📅 Carta Gantt":
    st.markdown('<div class="section">CARTA GANTT · PROGRAMA OFICIAL SAN FRANCISCO 211</div>', unsafe_allow_html=True)
    st.caption("Alternativa 2: cada piso termina el último día hábil del mismo mes en que inicia. El progreso se actualiza desde la base online.")

    cg1, cg2 = st.columns([2,1])
    with cg1:
        phase_filter = st.selectbox(
            "Mostrar", ["Todas las fases"] + [p.replace("FASE","Fase ") for p in allowed_phases()],
            key="gantt_phase_filter"
        )
    with cg2:
        try:
            from zoneinfo import ZoneInfo
            _today_cl = pd.Timestamp.now(tz=ZoneInfo("America/Santiago")).date()
        except Exception:
            _today_cl = pd.Timestamp.today().date()
        review_date = st.date_input("Fecha de revisión", value=_today_cl, key="gantt_review_date")

    selected = allowed_phases() if phase_filter == "Todas las fases" else [phase_filter.upper().replace("FASE ","FASE")]
    gdf = build_gantt_df(st.session_state.workbook_path, selected, review_date=review_date)

    if gdf.empty:
        st.warning("No hay datos disponibles para construir la Carta Gantt.")
    else:
        phase_rank={'Fase 1':1,'Fase 2':2,'Fase 3':3,'Fase 4':4}
        gdf['Piso_num']=gdf['Piso'].str.extract(r'(\d+)').astype(int)
        gdf['Fase_num']=gdf['Fase'].map(phase_rank).fillna(99).astype(int)
        gdf=gdf.sort_values(['Fase_num','Piso_num'],kind='stable').reset_index(drop=True)
        gdf['Etiqueta']=gdf.apply(lambda r: f"{r['Fase']} · Piso {int(r['Piso_num'])} | {r['Inicio'].strftime('%d-%m-%Y')} → {r['Término'].strftime('%d-%m-%Y')}",axis=1)
        category_order=gdf['Etiqueta'].tolist()
        fig = px.timeline(
            gdf, x_start="Inicio", x_end="Término", y="Etiqueta", color="Fase", text="Progreso (%)",
            hover_data={
                "Piso":True,"Días":True,"Progreso (%)":':.1f',"Días atraso":True,"Estado plazo":True,
                "Inicio":'|%d-%m-%Y',"Término":'|%d-%m-%Y'
            }
        )
        fig.update_traces(texttemplate="%{text:.1f}%", textposition="inside")
        fig.update_yaxes(title=None, categoryorder='array', categoryarray=category_order, autorange="reversed", tickmode='array', tickvals=category_order, ticktext=category_order)
        fig.update_xaxes(title="Programación oficial", tickformat="%d-%m-%Y")
        # Línea roja = fecha desde la cual se calculan los días de atraso.
        fig.add_vline(x=pd.Timestamp(review_date).timestamp()*1000, line_dash="dash", line_color="red")
        fig.update_layout(height=max(520,30*len(gdf)+170), legend_title_text="Fase", margin=dict(l=10,r=10,t=30,b=10))
        st.plotly_chart(fig, use_container_width=True)

        delayed = int((gdf["Días atraso"] > 0).sum())
        max_delay = int(gdf["Días atraso"].max()) if not gdf.empty else 0
        a1,a2,a3 = st.columns(3)
        a1.metric("Fecha de revisión", pd.Timestamp(review_date).strftime("%d-%m-%Y"))
        a2.metric("Actividades atrasadas", delayed)
        a3.metric("Mayor atraso", f"{max_delay} días")

        detail = gdf.sort_values(["Fase_num","Piso_num"])[["Fase","Piso","Inicio","Término","Días","Progreso (%)","Días atraso","Estado plazo"]].copy()
        detail["Inicio"] = detail["Inicio"].dt.strftime("%d-%m-%Y")
        detail["Término"] = detail["Término"].dt.strftime("%d-%m-%Y")
        st.dataframe(style_percent_df(detail, ["Progreso (%)"]), use_container_width=True, hide_index=True, height=420)

elif page == "⚠️ Ruta crítica":
    st.markdown("""
    <style>
    .route-head{display:flex;align-items:center;justify-content:space-between;background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px 16px;margin-bottom:10px}
    .route-head-title{font-size:24px;font-weight:900;color:#14263e}.route-head-sub{font-size:12px;color:#667085;margin-top:2px}
    .route-kpi{background:#fff;border:1px solid #dfe6ef;border-radius:8px;padding:12px 14px;min-height:124px;box-shadow:0 2px 8px rgba(15,35,60,.04)}
    .route-kpi .t{font-size:12px;font-weight:850;text-transform:uppercase;color:#334155}.route-kpi .v{font-size:30px;font-weight:900;color:#0f172a;margin:8px 0 2px}.route-kpi .s{font-size:11px;color:#64748b}
    .route-panel{background:#fff;border:1px solid #dfe6ef;border-radius:8px;padding:10px 10px 12px;box-shadow:0 2px 8px rgba(15,35,60,.04)}
    .route-panel-title{font-size:15px;font-weight:900;color:#d01e1e;margin:0 0 8px 2px}
    .route-table-wrap{overflow:auto;max-height:820px;border:1px solid #d7dee8}.route-table{width:100%;border-collapse:collapse;font-size:11.5px;background:#fff}.route-table th{position:sticky;top:0;z-index:3;background:#f8fafc;color:#0f172a;border:1px solid #d7dee8;padding:9px 7px;font-size:10.5px}.route-table td{border:1px solid #e2e8f0;padding:7px 6px;vertical-align:middle}.phase-band{color:#fff;font-weight:900;text-align:center;width:54px}.phase-blank{color:transparent}.task-name{min-width:270px}.pct-cell{min-width:165px;white-space:nowrap}.mini-track{display:inline-block;width:100px;height:16px;background:#e5e7eb;border-radius:2px;margin-right:6px;vertical-align:middle;overflow:hidden}.mini-fill{height:100%}.state-cell{white-space:nowrap}.critical-cell{text-align:center;white-space:nowrap}
    </style>
    """, unsafe_allow_html=True)

    snap = db_latest_weekly_snapshot()
    if snap:
        route_general=float(snap.get("general", general) or 0)
        route_summaries={"FASE1":float(snap.get("fase1", summaries.get("FASE1",0)) or 0),"FASE2":float(snap.get("fase2", summaries.get("FASE2",0)) or 0),"FASE3":float(snap.get("fase3", summaries.get("FASE3",0)) or 0),"FASE4":float(snap.get("fase4", summaries.get("FASE4",0)) or 0)}
        route_date=pd.to_datetime(snap.get("update_date"), errors="coerce")
    else:
        route_general=general; route_summaries=summaries; route_date=official_snapshot_date
    date_txt = route_date.strftime('%d-%m-%Y') if route_date is not None and not pd.isna(route_date) else 'Sin corte oficial'
    st.markdown('<div class="route-head"><div><div class="route-head-title">🔗 RUTA CRÍTICA DEL PROYECTO</div><div class="route-head-sub">La ruta crítica y la Carta Gantt se actualizan con el avance real registrado.</div></div><div style="text-align:right;font-size:12px;color:#475569">📅 Fecha de actualización<br><b style="color:#d01e1e">%s</b></div></div>' % date_txt, unsafe_allow_html=True)

    kcols=st.columns([1.2,1,1,1,1,2.2])
    with kcols[0]:
        st.markdown('<div class="route-kpi"><div class="t">Avance general del proyecto</div><div class="v">%.1f%%</div><div class="s">Última actualización: <b>%s</b></div></div>' % (route_general,date_txt), unsafe_allow_html=True)
    for i,ph in enumerate(PHASES, start=1):
        with kcols[i]:
            color=_phase_hex(ph)
            st.markdown('<div class="route-kpi" style="border-top:7px solid %s"><div class="t">%s</div><div class="v">%.0f%%</div><div class="s">Avance real</div></div>' % (color,ph.replace('FASE','FASE '),route_summaries.get(ph,0)), unsafe_allow_html=True)
    with kcols[5]:
        hist=load_weekly_history(st.session_state.workbook_path)
        if not hist.empty:
            plot=hist.tail(8).copy(); cols=[c for c in ['Fase 1','Fase 2','Fase 3','Fase 4'] if c in plot.columns]
            long=plot.melt(id_vars=['Fecha actualización'], value_vars=cols, var_name='Fase', value_name='Avance')
            figw=px.line(long,x='Fecha actualización',y='Avance',color='Fase',markers=True)
            figw.update_layout(height=125,margin=dict(l=0,r=0,t=8,b=0),legend=dict(orientation='h',y=1.2,x=0),font=dict(size=9))
            figw.update_yaxes(title=None,ticksuffix='%'); figw.update_xaxes(title=None,tickformat='%d-%m')
            st.plotly_chart(figw,use_container_width=True,config={'displayModeBar':False})
        else: st.info('Sin historial semanal')

    selected_phases=allowed_phases()
    # Alternativa 2: paneles apilados, cada uno a todo el ancho disponible.
    st.markdown('<div class="route-panel"><div class="route-panel-title">RUTA CRÍTICA ACTUALIZADA</div>',unsafe_allow_html=True)
    rdf=route_activity_summary(st.session_state.workbook_path, selected_phases)
    st.markdown(render_route_activity_html(rdf), unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="route-panel"><div class="route-panel-title">CARTA GANTT – PROGRAMA ACTUALIZADO</div>',unsafe_allow_html=True)
    try:
        from zoneinfo import ZoneInfo
        route_review_date=pd.Timestamp.now(tz=ZoneInfo('America/Santiago')).date()
    except Exception: route_review_date=pd.Timestamp.today().date()
    gdf=build_gantt_df(st.session_state.workbook_path, selected_phases, review_date=route_review_date)
    if not gdf.empty:
        def gantt_state(row):
            p=float(row['Progreso (%)'])
            if p>=100: return 'Completado (100%)'
            if p>0 and row['Días atraso']<=0: return 'En progreso'
            if row['Días atraso']>0: return 'En riesgo'
            return 'No iniciado (0%)'
        gdf['Estado visual']=gdf.apply(gantt_state,axis=1)
        cmap={'Completado (100%)':'#1f8f3a','En progreso':'#7bc96f','En riesgo':'#f6bf26','No iniciado (0%)':'#c9c9c9'}
        # Orden fijo y correlativo: Fase 1 Piso 1-9, luego Fase 2, Fase 3 y Fase 4.
        phase_rank={'Fase 1':1,'Fase 2':2,'Fase 3':3,'Fase 4':4}
        gdf['Piso_num']=gdf['Piso'].str.extract(r'(\d+)').astype(int)
        gdf['Fase_num']=gdf['Fase'].map(phase_rank).fillna(99).astype(int)
        gdf=gdf.sort_values(['Fase_num','Piso_num'],kind='stable').reset_index(drop=True)
        gdf['Etiqueta']=gdf.apply(lambda r: f"{r['Fase']} · Piso {int(r['Piso_num'])}   |   {r['Inicio'].strftime('%d-%m-%Y')} → {r['Término'].strftime('%d-%m-%Y')}",axis=1)
        category_order=gdf['Etiqueta'].tolist()
        fig=px.timeline(gdf,x_start='Inicio',x_end='Término',y='Etiqueta',color='Estado visual',text='Progreso (%)',color_discrete_map=cmap,hover_data={'Fase':True,'Piso':True,'Inicio':'|%d-%m-%Y','Término':'|%d-%m-%Y','Días':True,'Días atraso':True,'Estado plazo':True})
        fig.update_traces(texttemplate='%{text:.0f}%',textposition='outside')
        fig.update_yaxes(title=None,tickfont=dict(size=9),categoryorder='array',categoryarray=category_order,autorange='reversed',tickmode='array',tickvals=category_order,ticktext=category_order)
        fig.update_xaxes(title=None,tickformat='%d-%m-%Y',side='top',gridcolor='#e5e7eb')
        fig.add_vline(x=pd.Timestamp(route_review_date).timestamp()*1000,line_color='#e11d1d',line_dash='dash',line_width=1.5)
        fig.update_layout(height=max(980, 27*len(gdf)+90),margin=dict(l=0,r=8,t=45,b=8),legend=dict(orientation='h',y=-.07,x=0,font=dict(size=9)),plot_bgcolor='white',paper_bgcolor='white')
        st.plotly_chart(fig,use_container_width=True,config={'displayModeBar':False})
        delayed=int((gdf['Días atraso']>0).sum()); max_delay=int(gdf['Días atraso'].max()) if len(gdf) else 0
        m1,m2,m3=st.columns(3); m1.metric('HOY / revisión',pd.Timestamp(route_review_date).strftime('%d-%m-%Y')); m2.metric('Actividades atrasadas',delayed); m3.metric('Mayor atraso',f'{max_delay} días')
    else: st.warning('Sin datos para la Carta Gantt.')
    st.markdown('</div>',unsafe_allow_html=True)

elif page == "⬆️ Importar / Exportar":
    if not can_access_admin_tools():
        st.error("Solo el Administrador puede importar o exportar archivos.")
        st.stop()
    st.info(
        "La descarga usa siempre la plantilla Excel depurada de esta versión. "
        "AVANCE SEMANAL 1 queda enlazado a las partidas de FASE1–FASE4."
    )

    uploaded=st.file_uploader("Cargar una nueva versión de CONTROL_FASES_SFCO211.xlsx",type=["xlsx"])
    if uploaded is not None:
        # Evita reprocesar el mismo archivo en cada rerun de Streamlit.
        _uploaded_bytes = uploaded.getvalue()
        _uploaded_signature = f"{uploaded.name}:{len(_uploaded_bytes)}:{hash(_uploaded_bytes)}"
        if st.session_state.get("last_uploaded_workbook_signature") != _uploaded_signature:
            with open(st.session_state.workbook_path,"wb") as f:
                f.write(_uploaded_bytes)
            # get_sheet_names sí usa st.cache_data; load_phase NO está cacheada.
            get_sheet_names.clear()
            st.session_state["last_uploaded_workbook_signature"] = _uploaded_signature
            st.session_state["uploaded_workbook_ok"] = True
            st.success("✅ Archivo cargado correctamente. Ya puedes confirmar la actualización hacia Supabase.")
            st.rerun()
        elif st.session_state.get("uploaded_workbook_ok"):
            st.success("✅ Archivo cargado correctamente. Ya puedes confirmar la actualización hacia Supabase.")

    st.markdown("### 🔄 Convertir este Excel en la nueva base oficial de Supabase")
    st.caption(
        "Esta función toma TODAS las partidas de FASE1, FASE2, FASE3 y FASE4 del Excel actual. "
        "Los valores del Excel reemplazan los porcentajes de las mismas celdas en phase_updates. "
        "Se usa upsert por Fase + fila Excel + partida, por lo que no crea duplicados."
    )

    # El control de seguridad se muestra SIEMPRE. La validación pesada del Excel
    # se ejecuta solo cuando el Administrador decide sincronizar.
    st.warning("Antes de continuar, verifica que el archivo cargado sea el Excel de recuperación modificado que quieres dejar como base oficial.")
    _confirm_recovery = st.checkbox(
        "✅ Confirmo que quiero reemplazar los avances de Supabase con los porcentajes de ESTE Excel.",
        key="confirm_recovery_to_supabase_v47"
    )
    _sync_clicked = st.button(
        "🔄 ACTUALIZAR BASE SUPABASE CON ESTE EXCEL",
        type="primary",
        use_container_width=True,
        disabled=not _confirm_recovery,
        key="sync_recovery_to_supabase_v47",
    )
    st.caption("El proceso usa upsert: no suma porcentajes ni duplica registros. El valor del Excel pasa a ser el valor oficial de cada partida coincidente.")

    if _sync_clicked:
        try:
            with st.spinner("1/3 Validando FASE1–FASE4 del Excel…"):
                _preview_payload, _preview_counts = workbook_recovery_payload(
                    st.session_state.workbook_path,
                    st.session_state.get("user_name", "Administrador")
                )
            st.success(
                "Excel validado: "
                + " · ".join([f"{_p}: {_preview_counts.get(_p,0):,}".replace(",", ".") for _p in PHASES])
                + f" · Total: {len(_preview_payload):,}".replace(",", ".")
            )
            with st.spinner("2/3 Actualizando Supabase y 3/3 verificando los valores guardados…"):
                _ok, _msg, _counts = db_replace_from_recovery_workbook(
                    st.session_state.workbook_path,
                    st.session_state.get("user_name", "Administrador")
                )
            if _ok:
                st.session_state["recovery_sync_message"] = _msg
                st.success("✅ BASE SUPABASE ACTUALIZADA CORRECTAMENTE")
                st.success(_msg)
                st.info("Ahora presiona 'Sincronizar datos' arriba o vuelve al Dashboard para ver los porcentajes oficiales reconstruidos desde este Excel.")
            else:
                st.error(f"No se completó la sincronización: {_msg}")
        except Exception as _e:
            st.error(f"El Excel no pudo sincronizarse: {_e}")

    st.markdown("### Respaldo de base online")
    online_updates = db_phase_updates()
    if online_updates:
        backup_df = pd.DataFrame(online_updates)
        preferred = [c for c in ["phase","excel_row","activity","percent","updated_by","updated_at","id"] if c in backup_df.columns]
        other = [c for c in backup_df.columns if c not in preferred]
        backup_df = backup_df[preferred + other].sort_values(
            [c for c in ["updated_at","phase","excel_row"] if c in backup_df.columns],
            ascending=True
        )
        c1,c2,c3 = st.columns(3)
        c1.metric("Cambios guardados online", f"{len(backup_df):,}".replace(",","."))
        c2.metric("Departamentos con cambios", backup_df[[c for c in ["phase","excel_row"] if c in backup_df.columns]].drop_duplicates().shape[0])
        latest = str(backup_df["updated_at"].max()) if "updated_at" in backup_df.columns else "—"
        c3.metric("Última actualización", latest[:19].replace("T"," "))
        st.success("La base online tiene detalle de avances. Descarga este respaldo antes de importar o reemplazar cualquier Excel.")
        st.download_button(
            "🛟 DESCARGAR RESPALDO BASE ONLINE (CSV)",
            data=backup_df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"RESPALDO_BASE_ONLINE_SFCO211_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
        with st.expander("Ver últimos cambios guardados online"):
            st.dataframe(backup_df.tail(100), use_container_width=True, height=420)
    else:
        st.warning("No se encontraron cambios detallados en phase_updates. No importes otro Excel hasta revisar la conexión de Supabase.")

    st.markdown("### 🧭 Diagnóstico de recuperación histórica")
    st.caption(
        "Compara el detalle actualmente reconstruible desde phase_updates con el último corte semanal. "
        "Una diferencia positiva significa que el histórico confirma más avance del que puede reconstruirse celda por celda. "
        "La aplicación NO inventa qué partidas faltan."
    )
    _hist_latest = db_latest_weekly_snapshot()
    _live_diag = {p: phase_summary(st.session_state.workbook_path, p) for p in PHASES}
    if _hist_latest:
        _official_diag = {
            "FASE1": float(pd.to_numeric(_hist_latest.get("fase1"), errors="coerce") or 0),
            "FASE2": float(pd.to_numeric(_hist_latest.get("fase2"), errors="coerce") or 0),
            "FASE3": float(pd.to_numeric(_hist_latest.get("fase3"), errors="coerce") or 0),
            "FASE4": float(pd.to_numeric(_hist_latest.get("fase4"), errors="coerce") or 0),
        }
        _diag_rows=[]
        for _p in PHASES:
            _gap=round(_official_diag[_p]-float(_live_diag.get(_p,0)),1)
            _diag_rows.append({
                "Fase": _p.replace("FASE","Fase "),
                "Detalle recuperable (%)": round(float(_live_diag.get(_p,0)),1),
                "Último corte (%)": round(_official_diag[_p],1),
                "Brecha no reconstruible (pp)": max(0.0,_gap),
                "Estado": "✅ Completo" if _gap <= 0.05 else "⚠️ Falta detalle histórico",
            })
        _diag_df=pd.DataFrame(_diag_rows)
        st.dataframe(_diag_df, use_container_width=True, hide_index=True)
        _missing=_diag_df[_diag_df["Brecha no reconstruible (pp)"]>0.05]
        if len(_missing):
            st.warning(
                "El último corte semanal confirma un avance mayor que el detalle actualmente disponible en Supabase. "
                "Con solo los porcentajes globales no es matemáticamente posible saber qué departamentos/partidas faltan. "
                "No se realizará ninguna reconstrucción automática que pueda inventar datos."
            )
        else:
            st.success("El detalle online disponible alcanza o supera el último corte semanal en todas las fases.")
    else:
        st.info("No existe un corte semanal con el cual comparar el detalle online.")

    _audit_rows = db_phase_update_history()
    if _audit_rows:
        _audit_df=pd.DataFrame(_audit_rows)
        st.success(f"Historial detallado v41 activo: {len(_audit_df):,} cambios auditables.".replace(",","."))
        st.download_button(
            "🧾 Descargar historial detallado de cambios (CSV)",
            data=_audit_df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"HISTORIAL_CAMBIOS_SFCO211_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    else:
        st.info(
            "El historial inmutable empieza a poblarse desde la v41. Si aún no creaste las tablas nuevas en Supabase, "
            "ejecuta el archivo supabase_setup.sql incluido en esta versión. El autoguardado actual seguirá funcionando igual."
        )

    st.markdown("### Exportar Excel recuperado")
    st.caption(
        "Esta descarga parte de la plantilla oficial y aplica TODOS los cambios detallados que actualmente existen en Supabase. "
        "Úsala como respaldo recuperado del estado online."
    )

    export_data = build_formatted_export()

    st.download_button(
        "⬇️ Descargar Excel RECUPERADO desde base online",
        data=export_data,
        file_name="CONTROL_FASES_SFCO211_RECUPERADO_BASE_ONLINE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    if st.button("Restablecer archivo original"):
        shutil.copy2(BASE,st.session_state.workbook_path)
        load_phase.clear(); get_sheet_names.clear()
        st.rerun()
